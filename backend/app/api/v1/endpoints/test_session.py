"""Test Session CRUD endpoints."""

import base64
import logging
import math
from datetime import date
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select as sa_select
from sqlalchemy.orm import Session

from app.crud.lookup import DuplicateError
from app.crud.student import get_students_by_session_and_zone
from app.crud.test_session import (
    STATE_KEY_ACTIVE,
    STATE_KEY_EMBEDDING,
    STATE_KEY_LOADING,
    add_smena_to_session,
    change_session_state,
    count_students_by_session_and_zone,
    count_students_per_smena_by_zone,
    create_test_session,
    delete_test_session,
    get_active_test_sessions,
    get_session_states,
    get_smenas,
    get_test_session,
    get_test_sessions_paginated,
    get_tests,
    remove_smena_from_session,
    update_test_session,
)
from app.core.permissions import P
from app.dependencies import PermissionChecker, get_current_active_user, get_db
from app.models.session_state import SessionState
from app.models.test_session import TestSession
from app.models.user import User
from app.models.zone import Zone
from app.schemas.student import (
    PassportUpdateRequest,
    PassportUpdateResult,
    StudentResponse,
)
from app.services.passport_updater import (
    parse_passport_excel,
    update_session_passports,
)
from app.schemas.test_session import (
    ActiveSmenaResponse,
    ActiveTestSessionResponse,
    SessionStateResponse,
    SmenaResponse,
    TestResponse,
    TestSessionCreate,
    TestSessionListResponse,
    TestSessionResponse,
    TestSessionSmenaCreate,
    TestSessionSmenaResponse,
    TestSessionUpdate,
)
from app.models.test_session_smena import TestSessionSmena
from app.schemas.dashboard_stats import DashboardStatsResponse
from app.services.embedding_extractor import get_embedding_progress
from app.services.session_dashboard_stats import (
    DashboardStatsError,
    get_dashboard_stats,
)
from app.services.session_stats_excel import build_session_stats_excel
from app.services.student_loader import get_student_load_progress
from app.tasks.excel_loader_task import excel_load_and_enrich_task
from app.tasks.student_loader_task import load_students_task
from app.tasks.verify_task import process_embeddings, process_retry_embeddings

# Maksimal yuklanadigan Excel hajmi (16 MB) — bu 100k qatordan ortiq Excel'lar
# uchun ham yetarli. Ko'paytirish kerak bo'lsa shu konstantani o'zgartiring.
_MAX_EXCEL_BYTES = 16 * 1024 * 1024
_ALLOWED_EXCEL_MIME = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/octet-stream",  # ba'zi browserlar shu bilan yuborishi mumkin
}

logger = logging.getLogger(__name__)

router = APIRouter()


def _resolve_effective_zone(
    db: Session, current_user: User, requested_zone_id: int | None
) -> tuple[int, str]:
    """Client tomonidan uzatilgan `zone_id`ni tekshirib, effective
    (zone_id, zone_name) juftligini qaytaradi.

    Xatti-harakat:
    - `requested_zone_id` None bo'lsa — foydalanuvchining o'z zonasi ishlatiladi
      (eski xulq-atvor saqlanadi).
    - Aks holda — berilgan zona user region'iga tegishliligi tekshiriladi.
      Tegishli emas bo'lsa 403 qaytadi. Shu bilan bir userning boshqa
      region zonalariga ma'lumot so'rashi bloklanadi.
    """
    user_zone_id = current_user.zone_id
    user_region_id = current_user.region_id
    user_zone_name = current_user.zone_name or ""

    if requested_zone_id is None or requested_zone_id == user_zone_id:
        return int(user_zone_id) if user_zone_id else 0, user_zone_name

    if not user_region_id:
        raise HTTPException(
            status_code=400,
            detail="Foydalanuvchiga region biriktirilmagan",
        )

    zone = db.get(Zone, int(requested_zone_id))
    if not zone:
        raise HTTPException(status_code=404, detail="Zone topilmadi")
    if int(zone.region_id) != int(user_region_id):
        raise HTTPException(
            status_code=403,
            detail="Tanlangan zone foydalanuvchi regioniga tegishli emas",
        )
    return int(zone.id), zone.name


def _resolve_effective_region(
    db: Session, current_user: User, requested_region_id: int | None
) -> int:
    """Statistika olinadigan effective `region_id`ni aniqlaydi.

    Region-darajadagi endpointlar uchun — zonadan mustaqil.

    Xatti-harakat:
    - `requested_region_id` berilmasa — foydalanuvchining bevosita biriktirilgan
      `region_id`i ishlatiladi. Eski akkaunt (region_id to'ldirilmagan) bo'lsa,
      region zona orqali aniqlanadi.
    - Berilgan bo'lsa — user o'z regioniga tegishliligi tekshiriladi (boshqa
      region ma'lumotini so'rashi bloklanadi).
    - Hech qaysi yo'l bilan region topilmasa — 400 qaytadi.
    """
    user_region_id = current_user.region_id
    # Eski akkaunt: region_id hali to'ldirilmagan bo'lsa, zona orqali aniqlaymiz.
    if not user_region_id and current_user.zone_id:
        zone = db.get(Zone, int(current_user.zone_id))
        if zone:
            user_region_id = zone.region_id

    if requested_region_id is None:
        effective = user_region_id
    else:
        effective = int(requested_region_id)
        if user_region_id and effective != int(user_region_id):
            raise HTTPException(
                status_code=403,
                detail="Tanlangan region foydalanuvchiga tegishli emas",
            )

    if not effective:
        raise HTTPException(
            status_code=400,
            detail="Foydalanuvchiga region biriktirilmagan",
        )
    return int(effective)


# --- Справочники (lookups) ---


@router.get("/tests", response_model=list[TestResponse])
def list_tests(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_active_user),
):
    """Barcha faol testlar ro'yxati."""
    return get_tests(db)


@router.get("/smenas", response_model=list[SmenaResponse])
def list_smenas(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_active_user),
):
    """Barcha faol smenalar ro'yxati."""
    return get_smenas(db)


@router.get("/states", response_model=list[SessionStateResponse])
def list_session_states(
    db: Session = Depends(get_db),
    _: User = Depends(get_current_active_user),
):
    """Barcha faol sessiya holatlari."""
    return get_session_states(db)


# --- Excel shabloni: static GET — /{session_id} wildcard'idan OLDIN ---
# Aks holda FastAPI "excel-template" ni session_id (int) sifatida talqin
# qiladi va 422 "Kiritilgan ma'lumotlar noto'g'ri" qaytaradi.

_TEMPLATE_HEADERS: tuple[tuple[str, str], ...] = (
    ("last_name", "Familiya"),
    ("first_name", "Ism"),
    ("middle_name", "Otasining ismi"),
    ("imei", "JShShIR (14 ta raqam)"),
    ("ps_ser", "Pasport seriyasi"),
    ("ps_num", "Pasport raqami"),
    ("region_number", "Region raqami"),
    ("zone_number", "Zona raqami (ixtiyoriy)"),
    ("smena_number", "Smena raqami"),
    ("gr_n", "Guruh (ixtiyoriy)"),
    ("e_date", "Sana (YYYY-MM-DD)"),
    ("subject_name", "Fan (ixtiyoriy)"),
)
_TEMPLATE_REQUIRED = {
    "last_name", "first_name", "imei", "ps_ser", "ps_num",
    "region_number", "smena_number", "e_date",
}


def _build_excel_template() -> bytes:
    """Foydalanuvchi uchun bo'sh shablon: sarlavhalar + 2 ta misol qator."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Studentlar"

    required_fill = PatternFill("solid", fgColor="FFE7C2")  # och sariq
    optional_fill = PatternFill("solid", fgColor="E8F0FE")  # och ko'k
    header_font = Font(bold=True, color="1F2937")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (key, label) in enumerate(_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.alignment = center
        cell.fill = required_fill if key in _TEMPLATE_REQUIRED else optional_fill
        suffix = " (majburiy)" if key in _TEMPLATE_REQUIRED else " (ixtiyoriy)"
        cell.comment = Comment(label + suffix, "FaceID")

    examples = (
        ("ALIYEV", "AKMAL", "OYBEKOVICH", "32401200012345",
         "AA", "1234567", 1, 1, 1, 101, "2026-06-15", "MATEMATIKA"),
        ("KARIMOVA", "NODIRA", None, "33301200067890",
         "AB", "7654321", 1, None, 2, None, "2026-06-15", None),
    )
    for row_idx, row in enumerate(examples, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    widths = (14, 14, 18, 18, 8, 12, 10, 12, 10, 8, 14, 18)
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get(
    "/excel-template",
    summary="Excel shablonini yuklab olish",
    description=(
        "Studentlar ro'yxati uchun bo'sh shablon (.xlsx). "
        "Sarlavha + 2 ta misol qator. Sariq ustunlar — majburiy, "
        "ko'k ustunlar — ixtiyoriy."
    ),
)
def download_excel_template(
    _: User = Depends(PermissionChecker(P.TEST_SESSION_UPDATE.code)),
):
    content = _build_excel_template()
    headers = {
        "Content-Disposition": 'attachment; filename="students_template.xlsx"',
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        BytesIO(content),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers,
    )


def _build_passport_template() -> bytes:
    """Passport yangilash uchun bo'sh shablon: jshshir, ps_ser, ps_num."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Passportlar"

    headers = (
        ("jshshir", "JSHSHIR (14 ta raqam)"),
        ("ps_ser", "Pasport seriyasi"),
        ("ps_num", "Pasport raqami"),
    )
    fill = PatternFill("solid", fgColor="FFE7C2")
    header_font = Font(bold=True, color="1F2937")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (key, label) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.alignment = center
        cell.fill = fill
        cell.comment = Comment(label + " (majburiy)", "FaceID")

    examples = (
        ("32401200012345", "AA", "1234567"),
        ("33301200067890", "AB", "7654321"),
    )
    for row_idx, row in enumerate(examples, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, width in enumerate((20, 10, 14), start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get(
    "/passport-template",
    summary="Passport yangilash shablonini yuklab olish",
    description="Passport yangilash uchun bo'sh shablon (.xlsx): jshshir, ps_ser, ps_num.",
)
def download_passport_template(
    _: User = Depends(PermissionChecker(P.TEST_SESSION_UPDATE.code)),
):
    content = _build_passport_template()
    headers = {
        "Content-Disposition": 'attachment; filename="passport_template.xlsx"',
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        BytesIO(content),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers,
    )


# --- TestSession CRUD ---


@router.get("/active", response_model=list[ActiveTestSessionResponse])
def list_active_sessions(
    zone_id: int | None = Query(
        default=None,
        description=(
            "Ixtiyoriy: boshqa zonaga (bino) bog'liq statistikani olish uchun. "
            "Bo'sh bo'lsa — foydalanuvchining o'z zonasi ishlatiladi. "
            "Berilgan zona user region'iga tegishli bo'lishi shart."
        ),
    ),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Barcha aktiv test sessiyalar ro'yxati (is_active=True).

    Har bir sessiya uchun `zone_id` (yoki user zonasi) bo'yicha student soni
    va zona nomi qaytariladi. Desktop tomonidan region ichidagi har qanday
    zone tanlash uchun ishlatiladi.
    """
    sessions = get_active_test_sessions(db)
    effective_zone_id, effective_zone_name = _resolve_effective_zone(
        db, current_user, zone_id
    )

    result = []
    for session in sessions:
        data = ActiveTestSessionResponse.model_validate(session)
        data.zone_name = effective_zone_name
        if effective_zone_id:
            data.zone_student_count = count_students_by_session_and_zone(
                db, test_session_id=session.id, zone_id=effective_zone_id
            )
            smena_counts = count_students_per_smena_by_zone(
                db, test_session_id=session.id, zone_id=effective_zone_id
            )
            for smena in data.smenas:
                smena.sm_student_count = smena_counts.get(smena.id, 0)
        result.append(data)
    return result


@router.get("", response_model=TestSessionListResponse)
def list_sessions(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    is_active: bool | None = None,
    test_id: int | None = None,
    test_state_id: int | None = Query(
        default=None,
        description="SessionState.id bo'yicha filter (Yaratilgan, Yuklab olindi, Embedding, Tayyor, Yakunlangan)",
    ),
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_READ.code)),
):
    """Test sessiyalar ro'yxati (pagination bilan)."""
    items, total = get_test_sessions_paginated(
        db,
        page=page,
        per_page=per_page,
        is_active=is_active,
        test_id=test_id,
        test_state_id=test_state_id,
    )
    return TestSessionListResponse(
        items=items,
        total=total,
        page=page,
        per_page=per_page,
        pages=math.ceil(total / per_page) if total else 0,
    )


@router.get("/{session_id}/students", response_model=list[StudentResponse])
def list_session_students(
    session_id: int,
    zone_id: int | None = Query(
        default=None,
        description=(
            "Ixtiyoriy: boshqa zonaga (bino) bog'liq studentlarni olish uchun. "
            "Bo'sh bo'lsa — foydalanuvchining o'z zonasi ishlatiladi. "
            "Berilgan zona user region'iga tegishli bo'lishi shart."
        ),
    ),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Sessiyaga tegishli studentlar ro'yxati (tanlangan zona bo'yicha).

    Agar `zone_id` berilgan bo'lsa va u user region'iga tegishli bo'lsa —
    shu zona studentlarini qaytaradi. Aks holda user o'zining zonasidagi
    studentlarni oladi."""
    effective_zone_id, _name = _resolve_effective_zone(db, current_user, zone_id)
    if not effective_zone_id:
        raise HTTPException(
            status_code=400,
            detail="Foydalanuvchiga zona biriktirilmagan",
        )

    session = get_test_session(db, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Sessiya topilmadi")

    return get_students_by_session_and_zone(
        db, test_session_id=session_id, zone_id=effective_zone_id
    )


@router.get("/{session_id}", response_model=TestSessionResponse)
def get_session(
    session_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_READ.code)),
):
    """Bitta test sessiyani olish."""
    session = get_test_session(db, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Sessiya topilmadi")
    return session


@router.post("", response_model=TestSessionResponse, status_code=status.HTTP_201_CREATED)
def create_session(
    body: TestSessionCreate,
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_CREATE.code)),
):
    """Yangi test sessiya yaratish.

    - test tanlash
    - kunlar va smenalar belgilash
    """
    if body.finish_date < body.start_date:
        raise HTTPException(
            status_code=400,
            detail="Tugash sanasi boshlanish sanasidan oldin bo'lishi mumkin emas",
        )

    smenas_data = [s.model_dump() for s in body.smenas] if body.smenas else None

    try:
        session = create_test_session(
            db,
            test_id=body.test_id,
            name=body.name,
            start_date=body.start_date,
            finish_date=body.finish_date,
            count_sm_per_day=body.count_sm_per_day,
            smenas=smenas_data,
        )
    except DuplicateError as e:
        raise HTTPException(status_code=409, detail=e.message)
    return session


@router.patch("/{session_id}", response_model=TestSessionResponse)
def update_session(
    session_id: int,
    body: TestSessionUpdate,
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_UPDATE.code)),
):
    """Test sessiyani yangilash. Holatni o'zgartirish uchun /state endpointdan foydalaning."""
    data = body.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=400, detail="O'zgartirish uchun maydon berilmadi")

    try:
        session = update_test_session(db, session_id=session_id, data=data)
    except DuplicateError as e:
        raise HTTPException(status_code=409, detail=e.message)
    if not session:
        raise HTTPException(status_code=404, detail="Sessiya topilmadi")
    return session


class ChangeStateRequest(BaseModel):
    test_state_id: int


@router.patch("/{session_id}/state", response_model=TestSessionResponse)
def change_state(
    session_id: int,
    body: ChangeStateRequest,
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_UPDATE.code)),
):
    """Sessiya holatini o'zgartirish.

    - key=2 ga o'tganda tashqi API dan studentlar yuklanadi
    - key=4 ga o'tganda is_active=True bo'ladi

    Agar tashqi API xatolik bersa, holat eski holatiga qaytariladi.
    """
    # Eski holatni eslab qolish (rollback uchun)
    old_session = db.get(TestSession, session_id)
    if not old_session:
        raise HTTPException(status_code=404, detail="Sessiya topilmadi")
    previous_state_id = old_session.test_state_id

    try:
        session = change_session_state(
            db, session_id=session_id, new_state_id=body.test_state_id
        )
    except DuplicateError as e:
        raise HTTPException(status_code=409, detail=e.message)

    # key=2 → studentlarni yuklash (Celery task — fonda)
    # Sinxron yuklash 100k+ talaba uchun proxy timeout (504) beradi.
    # Frontend `/student-load-progress` endpoint orqali progress polling qiladi.
    new_state = db.get(SessionState, body.test_state_id)
    if new_state and new_state.key == STATE_KEY_LOADING:
        load_students_task.delay(session_id, previous_state_id)
        logger.info(
            "Session #%d: student loading Celery task boshlandi", session_id
        )

    # key=3 → face embedding chiqarish (Celery task)
    if new_state and new_state.key == STATE_KEY_EMBEDDING:
        process_embeddings.delay(session_id)
        logger.info("Session #%d: embedding Celery task boshlandi", session_id)

    # key=4 → sessiya faollashtirish — barcha studentlar is_ready=True bo'lishi kerak
    if new_state and new_state.key == STATE_KEY_ACTIVE:
        from sqlalchemy import func, select as sa_select
        from app.models.student import Student
        from app.models.test_session_smena import TestSessionSmena as TSSmena

        smena_ids = [
            row[0]
            for row in db.execute(
                sa_select(TSSmena.id).where(TSSmena.test_session_id == session_id)
            )
        ]
        not_ready_count = 0
        if smena_ids:
            not_ready_count = db.scalar(
                sa_select(func.count(Student.id)).where(
                    Student.session_smena_id.in_(smena_ids),
                    Student.is_ready.is_(False),
                )
            ) or 0

        if not_ready_count > 0:
            # Batafsil ma'lumot olish
            no_image_count = db.scalar(
                sa_select(func.count(Student.id)).where(
                    Student.session_smena_id.in_(smena_ids),
                    Student.is_image.is_(False),
                )
            ) or 0
            no_face_count = db.scalar(
                sa_select(func.count(Student.id)).where(
                    Student.session_smena_id.in_(smena_ids),
                    Student.is_image.is_(True),
                    Student.is_face.is_(False),
                )
            ) or 0

            parts: list[str] = []
            if no_image_count > 0:
                parts.append(f"{no_image_count} ta studentda rasm topilmadi")
            if no_face_count > 0:
                parts.append(f"{no_face_count} ta studentda yuz aniqlanmadi")
            remaining = not_ready_count - no_image_count - no_face_count
            if remaining > 0:
                parts.append(f"{remaining} ta student tayyor emas")
            detail = "Sessiyani faollashtirish mumkin emas: " + ", ".join(parts)
            _rollback_state(db, session_id, previous_state_id)
            raise HTTPException(status_code=400, detail=detail)

    return session


def _rollback_state(db: Session, session_id: int, previous_state_id: int) -> None:
    """Xatolik bo'lganda holatni eski holatiga qaytarish."""
    try:
        change_session_state(
            db,
            session_id=session_id,
            new_state_id=previous_state_id,
        )
    except Exception:
        logger.error(
            "Session #%d: holatni qaytarishda xatolik", session_id
        )


@router.get("/{session_id}/embedding-progress")
def embedding_progress(
    session_id: int,
    _: User = Depends(PermissionChecker(P.TEST_SESSION_READ.code)),
):
    """Embedding jarayoni progressini olish (Redis dan)."""
    progress = get_embedding_progress(session_id)
    if not progress:
        return {
            "current": 0, "total": 0, "success": 0,
            "no_image": 0, "no_face": 0, "errors": 0, "failed": 0,
            "percent": 0, "status": "idle",
        }
    return progress


@router.post(
    "/{session_id}/upload-excel",
    status_code=status.HTTP_202_ACCEPTED,
    summary="Excel'dan studentlarni yuklash (Celery background)",
)
async def upload_students_excel(
    session_id: int,
    file: UploadFile = File(..., description=".xlsx fayl"),
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_UPDATE.code)),
):
    """Test sessiyaga Excel orqali studentlarni yuklash.

    Talab:
    - Sessiya state.key=1 (yangi yaratilgan) bo'lishi shart.
    - Sessiyada kamida bitta smena bo'lishi kerak.
    - Fayl `.xlsx` formatida, hajmi <= 16MB.

    Endpoint darhol 202 qaytaradi va Celery background task'ni ishga tushiradi.
    Frontend `/student-load-progress` endpoint orqali progress polling qiladi.
    Yakunida sessiya state.key=2 ("Yuklab olindi") ga avtomatik o'tadi.
    """
    session = db.get(TestSession, session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Sessiya topilmadi")

    # Faqat "yaratilgan" (key=1) holatdagi sessiyalarga Excel yuklash mumkin.
    current_state = db.get(SessionState, session.test_state_id)
    if current_state is None or current_state.key != 1:
        raise HTTPException(
            status_code=400,
            detail=(
                "Excel yuklash faqat 'Yaratilgan' (yangi) holatdagi "
                "sessiyalar uchun mumkin"
            ),
        )

    # Sessiyada smena bormi?
    smena_count = db.scalar(
        sa_select(TestSessionSmena.id)
        .where(TestSessionSmena.test_session_id == session_id)
        .limit(1)
    )
    if smena_count is None:
        raise HTTPException(
            status_code=400,
            detail="Sessiyaga avval smena qo'shing — Excel'dagi qatorlarni "
            "shu smenalar bo'yicha bog'laymiz",
        )

    # MIME va kengaytma tekshiruvi (defense in depth — openpyxl o'zi ham tekshiradi)
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Fayl `.xlsx` formatida bo'lishi kerak"
        )
    if file.content_type and file.content_type not in _ALLOWED_EXCEL_MIME:
        # Faqat ogohlantirish — ba'zi browserlar generic MIME yuborishi mumkin
        logger.warning(
            "Excel upload: content_type=%s (kutilmagan, davom etilmoqda)",
            file.content_type,
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Fayl bo'sh")
    if len(content) > _MAX_EXCEL_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"Fayl hajmi {_MAX_EXCEL_BYTES // (1024 * 1024)}MB dan oshmasin",
        )

    encoded = base64.b64encode(content).decode("ascii")
    previous_state_id = int(session.test_state_id)

    task = excel_load_and_enrich_task.delay(
        session_id, encoded, previous_state_id
    )
    logger.info(
        "Excel upload Celery task'ga yuborildi: session=%d, task=%s, size=%d bytes",
        session_id,
        task.id,
        len(content),
    )
    return {
        "task_id": task.id,
        "status": "queued",
        "message": "Excel qayta ishlanmoqda — progress orqali kuzating",
    }


def _require_session(db: Session, session_id: int) -> TestSession:
    session = db.get(TestSession, session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Sessiya topilmadi")
    return session


@router.post(
    "/{session_id}/passport-update",
    response_model=PassportUpdateResult,
    summary="Passport (ps_ser/ps_num) ma'lumotlarini ommaviy yangilash (paste)",
)
def update_passports_json(
    session_id: int,
    payload: PassportUpdateRequest,
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_UPDATE.code)),
):
    """Excel'dan nusxalab qo'yilgan (paste) qatorlar bo'yicha passportlarni yangilash.

    `jshshir` (= talaba IMEI) bo'yicha shu sessiyadagi talaba topiladi va uning
    passport seriyasi/raqami yangilanadi. Jarayon sinxron — natija darhol qaytadi.
    """
    _require_session(db, session_id)
    rows = [r.model_dump() for r in payload.rows]
    return update_session_passports(db, session_id, rows)


@router.post(
    "/{session_id}/passport-update/excel",
    response_model=PassportUpdateResult,
    summary="Passport ma'lumotlarini Excel'dan ommaviy yangilash",
)
async def update_passports_excel(
    session_id: int,
    file: UploadFile = File(..., description=".xlsx fayl: jshshir, ps_ser, ps_num"),
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_UPDATE.code)),
):
    """`.xlsx` fayl (`jshshir, ps_ser, ps_num`) orqali passportlarni yangilash.

    Fayl serverda openpyxl bilan o'qiladi va `jshshir` bo'yicha shu sessiyadagi
    talabalar passporti yangilanadi. Jarayon sinxron — natija darhol qaytadi.
    """
    _require_session(db, session_id)

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Fayl `.xlsx` formatida bo'lishi kerak"
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Fayl bo'sh")
    if len(content) > _MAX_EXCEL_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"Fayl hajmi {_MAX_EXCEL_BYTES // (1024 * 1024)}MB dan oshmasin",
        )

    rows, errors = parse_passport_excel(content)
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors))
    if not rows:
        raise HTTPException(
            status_code=400, detail="Excel'da yangilanadigan qator topilmadi"
        )

    return update_session_passports(db, session_id, rows)


@router.get("/{session_id}/student-load-progress")
def student_load_progress(
    session_id: int,
    _: User = Depends(PermissionChecker(P.TEST_SESSION_READ.code)),
):
    """Tashqi API'dan studentlar yuklash jarayoni progressini olish (Redis'dan).

    Holatlar:
    - idle      — jarayon hali boshlanmagan yoki TTL tugagan
    - processing — yuklanmoqda
    - completed — muvaffaqiyatli tugadi
    - error     — xatolik bo'ldi (sessiya state'i avvalgisiga qaytarildi)
    """
    progress = get_student_load_progress(session_id)
    if not progress:
        return {
            "current": 0, "total": 0,
            "pages_done": 0, "pages_total": 0,
            "skipped": 0, "percent": 0,
            "status": "idle", "message": "",
        }
    return progress


@router.get(
    "/{session_id}/dashboard-stats",
    response_model=DashboardStatsResponse,
    summary="Sessiya statistika dashboard",
)
def session_dashboard_stats(
    session_id: int,
    scope: str = Query(
        "smena",
        description="Statistika ko'lami: smena | day | overall",
    ),
    session_smena_id: int | None = Query(
        None,
        description="TestSessionSmena.id — scope=smena uchun majburiy (kun+smena)",
    ),
    day: date | None = Query(
        None,
        description="Kun (YYYY-MM-DD) — scope=day uchun majburiy",
    ),
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.STATISTICS_READ.code)),
):
    """Tanlangan ko'lam (scope) uchun dashboard statistikasini qaytarish.

    - **scope=smena** — bitta kun+smena (`session_smena_id` majburiy).
    - **scope=day** — bitta kunning barcha smenalari (`day` majburiy).
    - **scope=overall** — sessiyaning barcha kun va smenalari.

    - **Real-time**: javobda `is_realtime=True` qaytsa (sessiya state.key=4),
      frontend bu endpoint'ni har necha sekundda qayta chaqirib turishi
      tavsiya etiladi.
    - **Boshqa holatlarda**: oxirgi ma'lumotni qaytaradi (cache emas, har safar
      DB'dan o'qiydi; agar polling kerak bo'lmasa frontend bir marta chaqiradi).

    `summary` — 4 ta dashboard cardini to'ldiradi.
    `regions` — region.number bo'yicha tartiblangan ro'yxat, har birida shu
    statistikalarning to'liq nusxasi.
    """
    try:
        return get_dashboard_stats(
            db,
            session_id=session_id,
            scope=scope,
            session_smena_id=session_smena_id,
            day=day,
        )
    except DashboardStatsError as e:
        raise HTTPException(status_code=400, detail=str(e))


def _safe_filename(name: str) -> str:
    """Fayl nomi uchun xavfsiz ASCII variant (Content-Disposition fallback)."""
    cleaned = "".join(
        ch if ch.isalnum() or ch in ("-", "_", " ") else "_" for ch in name
    ).strip().replace(" ", "_")
    return cleaned or "statistika"


@router.get(
    "/{session_id}/dashboard-stats/export",
    summary="Sessiya statistikasini Excel (.xlsx) hisobotiga eksport qilish",
)
def export_session_dashboard_stats(
    session_id: int,
    scope: str = Query(
        "smena",
        description="Statistika ko'lami: smena | day | overall",
    ),
    session_smena_id: int | None = Query(
        None,
        description="TestSessionSmena.id — scope=smena uchun majburiy (kun+smena)",
    ),
    day: date | None = Query(
        None,
        description="Kun (YYYY-MM-DD) — scope=day uchun majburiy",
    ),
    alphabet: str = Query(
        "cyrillic",
        description="Hisobot alifbosi: cyrillic (krill) | latin (o'zbek lotin)",
    ),
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.STATISTICS_READ.code)),
):
    """Tanlangan ko'lam (scope) statistikasini rasmiy "МАЪЛУМОТ" ko'rinishidagi
    Excel (.xlsx) hisobotiga yozib qaytaradi.

    Parametrlar `/dashboard-stats` bilan bir xil. Hisobot tepasida sessiya
    nomi, "МАЪЛУМОТ" sarlavhasi va o'ng burchakda sana/vaqt/smena yoziladi;
    so'ng har bir viloyat bo'yicha statistika jadvali va "Жами" yig'indisi.
    """
    session = db.get(TestSession, session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Sessiya topilmadi")

    try:
        stats = get_dashboard_stats(
            db,
            session_id=session_id,
            scope=scope,
            session_smena_id=session_smena_id,
            day=day,
        )
    except DashboardStatsError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Sarlavha: test sessiya nomi + test nomi (servis krillga o'tkazadi)
    session_name = session.name or ""
    test_name = session.test.name if session.test else ""
    title = " ".join(p for p in (session_name, test_name) if p) or f"Sessiya #{session_id}"

    # Sessiya kun oralig'i (umumiy/kunlik sarlavhalar uchun)
    session_days = db.execute(
        sa_select(TestSessionSmena.day).where(
            TestSessionSmena.test_session_id == session_id
        )
    ).scalars().all()
    day_from = min(session_days) if session_days else None
    day_to = max(session_days) if session_days else None

    latin = alphabet.strip().lower() == "latin"
    content = build_session_stats_excel(
        stats, title=title, day_from=day_from, day_to=day_to, latin=latin
    )

    # Fayl nomi: sessiya + kun (+ smena) — scope'ga qarab
    base = _safe_filename(session_name or f"sessiya_{session_id}")
    parts = [base]
    if stats.day is not None:
        parts.append(stats.day.isoformat())
    if stats.scope == "smena" and stats.smena_number is not None:
        parts.append(f"{stats.smena_number}-smena")
    elif stats.scope == "day":
        parts.append("kunlik")
    elif stats.scope == "overall":
        parts.append("umumiy")
    parts.append("lotin" if latin else "krill")
    filename = "_".join(parts) + ".xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        BytesIO(content),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers=headers,
    )


@router.get("/{session_id}/student-stats")
def session_student_stats(
    session_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_READ.code)),
):
    """Sessiya studentlari statistikasi: jami, tayyor, rasmsiz, yuzsiz."""
    from sqlalchemy import func, select as sa_select
    from app.models.student import Student
    from app.models.test_session_smena import TestSessionSmena as TSSmena

    smena_ids = [
        row[0]
        for row in db.execute(
            sa_select(TSSmena.id).where(TSSmena.test_session_id == session_id)
        )
    ]
    if not smena_ids:
        return {"total": 0, "ready": 0, "not_ready": 0, "no_image": 0, "no_face": 0}

    base = Student.session_smena_id.in_(smena_ids)
    total = db.scalar(sa_select(func.count(Student.id)).where(base)) or 0
    ready = db.scalar(
        sa_select(func.count(Student.id)).where(base, Student.is_ready.is_(True))
    ) or 0
    no_image = db.scalar(
        sa_select(func.count(Student.id)).where(base, Student.is_image.is_(False))
    ) or 0
    no_face = db.scalar(
        sa_select(func.count(Student.id)).where(
            base, Student.is_image.is_(True), Student.is_face.is_(False)
        )
    ) or 0

    return {
        "total": total,
        "ready": ready,
        "not_ready": total - ready,
        "no_image": no_image,
        "no_face": no_face,
    }


@router.get("/smenas/{smena_id}/attendance-stats")
def smena_attendance_stats(
    smena_id: int,
    zone_id: int | None = Query(
        default=None,
        description=(
            "Ixtiyoriy: tanlangan zona bo'yicha davomat statistikasini olish. "
            "Bo'sh bo'lsa — user o'z zonasi. Zona user region'iga tegishli bo'lishi shart."
        ),
    ),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Smena + bino (zone) kesimida davomat statistikasi.

    Tanlangan `zone_id` (yoki user o'zining zonasi) va berilgan smena bo'yicha
    jami, kirgan, kirmagan va chetlatilgan (cheating) sonini qaytaradi.
    """
    effective_zone_id, _name = _resolve_effective_zone(db, current_user, zone_id)
    if not effective_zone_id:
        raise HTTPException(
            status_code=400,
            detail="Foydalanuvchiga zona biriktirilmagan",
        )

    from sqlalchemy import func, select as sa_select
    from app.models.student import Student

    base = (
        (Student.session_smena_id == smena_id)
        & (Student.zone_id == effective_zone_id)
    )
    total = db.scalar(sa_select(func.count(Student.id)).where(base)) or 0
    entered = db.scalar(
        sa_select(func.count(Student.id)).where(base, Student.is_entered.is_(True))
    ) or 0
    cheating = db.scalar(
        sa_select(func.count(Student.id)).where(base, Student.is_cheating.is_(True))
    ) or 0

    return {
        "total": total,
        "entered": entered,
        "not_entered": max(0, total - entered),
        "cheating": cheating,
    }


@router.get("/smenas/{smena_id}/attendance-by-region")
def smena_attendance_by_region(
    smena_id: int,
    region_id: int | None = Query(
        default=None,
        description=(
            "Ixtiyoriy: statistika olinadigan region. Berilmasa — "
            "foydalanuvchining o'z regioni ishlatiladi. Statistika "
            "zonaga emas, regionga bog'liq."
        ),
    ),
    zone_id: int | None = Query(
        default=None,
        description=(
            "Ixtiyoriy: aktiv bino — faqat modal'da `is_active` highlight "
            "uchun. Statistika natijasiga ta'sir qilmaydi."
        ),
    ),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Smena + region kesimida har bir bino bo'yicha davomat statistikasi.

    Statistika foydalanuvchining (yoki so'ralgan) **regioni** bo'yicha
    aniqlanadi — zona biriktirilmagan foydalanuvchilar uchun ham ishlaydi.

    Qaytaradi:
      - test_day, smena_number, smena_name (test_session_smena dan);
      - region_id, region_name;
      - active_zone_id / active_zone_name (highlight uchun, bo'lishi shart emas);
      - zones: shu region'dagi har bino bo'yicha
        {zone_id, zone_name, zone_number, total, entered, not_entered, cheating, is_active}.

    Bitta GROUP BY query orqali samarali ravishda olinadi.
    """
    from sqlalchemy import case, func, select as sa_select
    from app.models.region import Region
    from app.models.smena import Smena
    from app.models.student import Student
    from app.models.test_session_smena import TestSessionSmena

    effective_region_id = _resolve_effective_region(db, current_user, region_id)
    # Highlight uchun aktiv bino — so'ralgan `zone_id` yoki user zonasi.
    # Bo'lmasligi mumkin; statistikaga ta'sir qilmaydi.
    active_zone_id = int(zone_id) if zone_id else (current_user.zone_id or None)

    tss_row = db.execute(
        sa_select(
            TestSessionSmena.id,
            TestSessionSmena.day,
            TestSessionSmena.number,
            Smena.name,
        )
        .join(Smena, Smena.id == TestSessionSmena.test_smena_id)
        .where(TestSessionSmena.id == smena_id)
    ).first()
    if not tss_row:
        raise HTTPException(status_code=404, detail="Smena topilmadi")

    region_row = db.execute(
        sa_select(Region.id, Region.name).where(Region.id == effective_region_id)
    ).first()
    if not region_row:
        raise HTTPException(status_code=404, detail="Region topilmadi")

    zones_in_region = db.execute(
        sa_select(Zone.id, Zone.name, Zone.number)
        .where(Zone.region_id == effective_region_id, Zone.is_active.is_(True))
        .order_by(Zone.number, Zone.name)
    ).all()

    zone_ids = [z.id for z in zones_in_region]
    stats_by_zone: dict[int, dict] = {
        z.id: {"total": 0, "entered": 0, "cheating": 0} for z in zones_in_region
    }

    if zone_ids:
        rows = db.execute(
            sa_select(
                Student.zone_id,
                func.count(Student.id).label("total"),
                func.sum(case((Student.is_entered.is_(True), 1), else_=0)).label(
                    "entered"
                ),
                func.sum(case((Student.is_cheating.is_(True), 1), else_=0)).label(
                    "cheating"
                ),
            )
            .where(
                Student.session_smena_id == smena_id,
                Student.zone_id.in_(zone_ids),
            )
            .group_by(Student.zone_id)
        ).all()
        for row in rows:
            stats_by_zone[row.zone_id] = {
                "total": int(row.total or 0),
                "entered": int(row.entered or 0),
                "cheating": int(row.cheating or 0),
            }

    # Aktiv bino faqat shu region ichida bo'lsagina highlight qilinadi.
    active_zone_name = ""
    if active_zone_id is not None:
        match = next((z for z in zones_in_region if z.id == active_zone_id), None)
        if match is None:
            # So'ralgan/user zonasi bu regionga tegishli emas — highlight yo'q.
            active_zone_id = None
        else:
            active_zone_name = match.name

    zones_payload = []
    for z in zones_in_region:
        s = stats_by_zone.get(z.id, {"total": 0, "entered": 0, "cheating": 0})
        total = s["total"]
        entered = s["entered"]
        zones_payload.append(
            {
                "zone_id": z.id,
                "zone_name": z.name,
                "zone_number": z.number,
                "total": total,
                "entered": entered,
                "not_entered": max(0, total - entered),
                "cheating": s["cheating"],
                "is_active": z.id == active_zone_id,
            }
        )

    return {
        "test_day": tss_row.day.isoformat() if tss_row.day else None,
        "smena_number": tss_row.number,
        "smena_name": tss_row.name,
        "region_id": region_row.id,
        "region_name": region_row.name,
        "active_zone_id": active_zone_id,
        "active_zone_name": active_zone_name,
        "zones": zones_payload,
    }


@router.post("/{session_id}/retry-embedding")
def retry_embedding(
    session_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_UPDATE.code)),
):
    """Faqat is_ready=False bo'lgan studentlar uchun qayta embedding olish.

    Celery task ga yuboriladi.
    """
    session = get_test_session(db, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Sessiya topilmadi")

    # is_ready=False studentlar borligini tekshirish
    from sqlalchemy import func, select as sa_select
    from app.models.student import Student
    from app.models.test_session_smena import TestSessionSmena as TSSmena

    smena_ids = [
        row[0]
        for row in db.execute(
            sa_select(TSSmena.id).where(TSSmena.test_session_id == session_id)
        )
    ]
    not_ready = 0
    if smena_ids:
        not_ready = db.scalar(
            sa_select(func.count(Student.id)).where(
                Student.session_smena_id.in_(smena_ids),
                Student.is_ready.is_(False),
            )
        ) or 0

    if not_ready == 0:
        raise HTTPException(
            status_code=400,
            detail="Barcha studentlar tayyor — qayta embedding kerak emas",
        )

    process_retry_embeddings.delay(session_id)
    logger.info("Session #%d: qayta embedding Celery task boshlandi (%d student)", session_id, not_ready)

    return {"message": f"{not_ready} ta student uchun qayta embedding boshlandi", "not_ready": not_ready}


@router.delete("/{session_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_session(
    session_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_DELETE.code)),
):
    """Test sessiyani o'chirish."""
    try:
        if not delete_test_session(db, session_id):
            raise HTTPException(status_code=404, detail="Sessiya topilmadi")
    except DuplicateError as e:
        raise HTTPException(status_code=409, detail=e.message)


# --- Smena management ---


@router.post(
    "/{session_id}/smenas",
    response_model=TestSessionSmenaResponse,
    status_code=status.HTTP_201_CREATED,
)
def add_smena(
    session_id: int,
    body: TestSessionSmenaCreate,
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_UPDATE.code)),
):
    """Sessiyaga smena qo'shish."""
    session = get_test_session(db, session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Sessiya topilmadi")

    try:
        smena = add_smena_to_session(
            db,
            test_session_id=session_id,
            test_smena_id=body.test_smena_id,
            day=body.day,
        )
    except DuplicateError as e:
        raise HTTPException(status_code=409, detail=e.message)
    return smena


@router.delete("/{session_id}/smenas/{smena_id}", status_code=status.HTTP_204_NO_CONTENT)
def remove_smena(
    session_id: int,
    smena_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(PermissionChecker(P.TEST_SESSION_UPDATE.code)),
):
    """Sessiyadan smena olib tashlash."""
    if not remove_smena_from_session(db, smena_id):
        raise HTTPException(status_code=404, detail="Smena topilmadi")
