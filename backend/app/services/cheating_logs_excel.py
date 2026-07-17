"""Chetlatilganlar (cheating logs) ro'yxatini Excel (.xlsx) ga eksport qilish.

Adminka "Chetlatilganlar" sahifasidagi jadval bilan bir xil ustunlar, joriy
filtrlar bo'yicha. Material uslubidagi hisobot ko'rinishida tuziladi:

  1-qator   — rangli sarlavha bandi (indigo fon, oq matn, markazda)
  2-qator   — yaratilgan sana/vaqt + jami soni (ochiq indigo fon)
  4-qator   — jadval sarlavhasi (indigo, oq matn, AutoFilter dropdownlari)
  5+        — ma'lumot qatorlari (zebra, nozik chegara, markazlashtirilgan)

Qulayliklar: sarlavha muzlatilgan (freeze), ustunlarda filtr (AutoFilter),
chop etishda sarlavha har betda takrorlanadi.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# O'zbekiston vaqti (UZT) — UTC+5, DST yo'q.
UZ_TZ = timezone(timedelta(hours=5))

_FONT_NAME = "Calibri"

# Material uslubidagi ranglar (ARGB, FF = to'liq alfa).
_C_TITLE_BG = "FF303F9E"        # indigo 700 — sarlavha bandi
_C_TITLE_FG = "FFFFFFFF"        # oq
_C_SUB_BG = "FFE8EAF6"          # indigo 50 — subtitr fon
_C_SUB_FG = "FF3949AB"          # indigo 600 — subtitr matn
_C_HEADER_BG = "FF3949AB"       # indigo 600 — jadval sarlavhasi
_C_HEADER_FG = "FFFFFFFF"       # oq
_C_ZEBRA_BG = "FFF4F5FB"        # juda ochiq indigo — juft qatorlar
_C_BORDER = "FFD6D9E6"          # nozik chegara (ko'kimtir kulrang)
_C_TEXT = "FF263238"            # asosiy matn — quyuq
_C_MUTED = "FF78909C"           # ikkilamchi matn (sana, vakil)
_C_REASON_FG = "FFC62828"       # sabab — qizil (chetlatish urg'usi)

_thin = Side(style="thin", color=_C_BORDER)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# (sarlavha, item-kaliti, kenglik, hizalash)
_COLUMNS: tuple[tuple[str, str, int, str], ...] = (
    ("№", "_idx", 5, "center"),
    ("F.I.SH.", "student_full_name", 30, "left"),
    ("JSHSHIR", "imei", 17, "center"),
    ("Sabab turi", "rejection_type", 18, "center"),
    ("Sababi", "rejection_reason", 30, "left"),
    ("Test", "test_name", 24, "left"),
    ("Viloyat", "region_name", 20, "left"),
    ("Bino", "zone_name", 20, "left"),
    ("Sana", "smena_date", 13, "center"),
    ("Smena", "smena_name", 12, "center"),
    ("Chetlatilgan vaqti", "rejected_at", 19, "center"),
    ("Vakil", "username", 16, "left"),
)


def _fmt_dt(value) -> str:
    """`datetime`ni O'zbekiston vaqtida 'YYYY-MM-DD HH:MM' ko'rinishida beradi.
    Naive vaqt UTC deb qabul qilinadi (prod Postgres Etc/UTC)."""
    if value is None:
        return "—"
    if isinstance(value, datetime):
        dt = value
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(UZ_TZ).strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _cell_text(item: dict, key: str, idx: int) -> str:
    if key == "_idx":
        return str(idx)
    val = item.get(key)
    if key in ("rejected_at", "smena_date"):
        return _fmt_dt(val)
    if val is None or val == "":
        return "—"
    return str(val)


def build_cheating_logs_excel(
    items: list[dict],
    *,
    title: str,
    generated_at: datetime | None = None,
) -> bytes:
    """Chetlatilganlar ro'yxatidan Material uslubidagi .xlsx bytes yig'adi.

    `items` — `get_cheating_logs_paginated` qaytaradigan lug'atlar ro'yxati.
    `title` — hisobot sarlavhasi. `generated_at` — hisobot yaratilgan vaqt
    (berilmasa hozirgi UZ vaqti)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Chetlatilganlar"
    ws.sheet_view.showGridLines = False

    n_cols = len(_COLUMNS)
    last_letter = get_column_letter(n_cols)

    # Ustun kengliklari
    for i, (_, _, width, _align) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    center = Alignment(horizontal="center", vertical="center")
    left_mid = Alignment(horizontal="left", vertical="center")

    # ── 1-qator: rangli sarlavha bandi ──
    ws.merge_cells(f"A1:{last_letter}1")
    tcell = ws.cell(row=1, column=1, value=title.upper())
    tcell.font = Font(name=_FONT_NAME, size=15, bold=True, color=_C_TITLE_FG)
    tcell.alignment = center
    for i in range(1, n_cols + 1):
        ws.cell(row=1, column=i).fill = PatternFill("solid", fgColor=_C_TITLE_BG)
    ws.row_dimensions[1].height = 34

    # ── 2-qator: yaratilgan vaqt + jami soni ──
    gen = generated_at or datetime.now(UZ_TZ)
    if gen.tzinfo is None:
        gen = gen.replace(tzinfo=timezone.utc).astimezone(UZ_TZ)
    else:
        gen = gen.astimezone(UZ_TZ)
    ws.merge_cells(f"A2:{last_letter}2")
    sub = ws.cell(
        row=2,
        column=1,
        value=(
            f"Yaratilgan: {gen.strftime('%Y-%m-%d %H:%M')}      "
            f"Jami: {len(items)} ta"
        ),
    )
    sub.font = Font(name=_FONT_NAME, size=10, bold=True, color=_C_SUB_FG)
    sub.alignment = center
    for i in range(1, n_cols + 1):
        ws.cell(row=2, column=i).fill = PatternFill("solid", fgColor=_C_SUB_BG)
    ws.row_dimensions[2].height = 22

    # 3-qator: nozik bo'shliq
    ws.row_dimensions[3].height = 6

    # ── 4-qator: jadval sarlavhasi ──
    header_row = 4
    header_fill = PatternFill("solid", fgColor=_C_HEADER_BG)
    header_font = Font(name=_FONT_NAME, size=11, bold=True, color=_C_HEADER_FG)
    header_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    for i, (label, _key, _w, _align) in enumerate(_COLUMNS, start=1):
        c = ws.cell(row=header_row, column=i, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = _border
    ws.row_dimensions[header_row].height = 34

    # ── Ma'lumot qatorlari ──
    zebra = PatternFill("solid", fgColor=_C_ZEBRA_BG)
    data_font = Font(name=_FONT_NAME, size=10, color=_C_TEXT)
    muted_font = Font(name=_FONT_NAME, size=9.5, color=_C_MUTED)
    reason_font = Font(name=_FONT_NAME, size=10, bold=True, color=_C_REASON_FG)

    row = header_row + 1
    for n, item in enumerate(items, start=1):
        for i, (_label, key, _w, align) in enumerate(_COLUMNS, start=1):
            c = ws.cell(row=row, column=i, value=_cell_text(item, key, n))
            if key == "rejection_reason":
                c.font = reason_font
            elif key in ("rejected_at", "smena_date", "username"):
                c.font = muted_font
            else:
                c.font = data_font
            c.alignment = left_mid if align == "left" else center
            if align == "left":
                c.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            c.border = _border
            if n % 2 == 0:
                c.fill = zebra
        ws.row_dimensions[row].height = 21
        row += 1

    last_data_row = row - 1

    # Sarlavha qatorida filtr (dropdown) — ustunlar bo'yicha saralash/filtrlash.
    if last_data_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{last_letter}{last_data_row}"

    # Sarlavha qatorini muzlatish (scroll'da tepada qoladi).
    ws.freeze_panes = f"A{header_row + 1}"

    # Chop etish: landshaft, gorizontal sig'dirish, sarlavha har betda takror.
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"1:{header_row}"
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
