"""Davomat bot — kelmaganlar ro'yxati Excel generator.

Tanlangan kontekstda — bitta smena, bitta kun (barcha smenalar) yoki butun
sessiya (barcha kun va smenalar) — foydalanuvchining biriktirilgan regionlar
kesimida `is_entered=False` va `is_applied=False` bo'lgan talabgorlarni
topadi va openpyxl orqali chiroyli formatlangan .xlsx hosil qiladi.

Tartib: kun → smena → viloyat → bino → guruh → joy.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.region import Region
from app.models.smena import Smena
from app.models.student import Student
from app.models.test_session import TestSession
from app.models.test_session_smena import TestSessionSmena
from app.models.zone import Zone


class AbsenteesError(Exception):
    """Smena topilmadi yoki ruxsat etilgan region yo'q."""


# Excel uslublari — bitta joyda, qayta foydalanish uchun.
_TITLE_FILL = PatternFill("solid", fgColor="1B5E20")  # to'q yashil (title)
_META_FILL = PatternFill("solid", fgColor="2E7D32")  # yashil (sana/smena)
_HEADER_FILL = PatternFill("solid", fgColor="2E7D32")  # yashil (jadval header)
_ZEBRA_FILL = PatternFill("solid", fgColor="E8F5E9")  # yashilning ochiq tovi (zebra)

_THIN = Side(style="thin", color="A5D6A7")
_THICK = Side(style="medium", color="1B5E20")

_BORDER_CELL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_HEADER = Border(left=_THIN, right=_THIN, top=_THICK, bottom=_THICK)


def _safe(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _resolve_smenas(
    db: Session,
    *,
    session_id: int,
    session_smena_id: int | None,
    test_day: date | None,
) -> tuple[TestSession, list[int], date | None, str, int, str]:
    """Excel uchun smenalar to'plamini aniqlaydi.

    Qaytaradi: (test_session, smena_ids, resolved_day, smena_name,
    smena_number, scope).
    """
    test_session = db.get(TestSession, session_id)
    if not test_session:
        raise AbsenteesError("Test sessiyasi topilmadi")

    if session_smena_id is not None:
        tss = db.get(TestSessionSmena, session_smena_id)
        if not tss:
            raise AbsenteesError("Smena topilmadi")
        if int(tss.test_session_id) != int(session_id):
            raise AbsenteesError("Tanlangan smena bu sessiyaga tegishli emas")
        smena = db.get(Smena, tss.test_smena_id)
        smena_name = smena.name if smena else ""
        smena_number = int(tss.number or (smena.number if smena else 0))
        return test_session, [int(tss.id)], tss.day, smena_name, smena_number, "smena"

    if test_day is not None:
        rows = db.execute(
            select(TestSessionSmena.id)
            .where(
                TestSessionSmena.test_session_id == session_id,
                TestSessionSmena.day == test_day,
                TestSessionSmena.is_active.is_(True),
            )
        ).all()
        ids = [int(r.id) for r in rows]
        if not ids:
            raise AbsenteesError("Tanlangan kunda aktiv smenalar topilmadi")
        return test_session, ids, test_day, "Kun yakuni", 0, "day"

    rows = db.execute(
        select(TestSessionSmena.id)
        .where(
            TestSessionSmena.test_session_id == session_id,
            TestSessionSmena.is_active.is_(True),
        )
    ).all()
    ids = [int(r.id) for r in rows]
    if not ids:
        raise AbsenteesError("Sessiyada aktiv smenalar topilmadi")
    return test_session, ids, None, "Umumiy", 0, "total"


def _fetch_absentees(
    db: Session, *, session_smena_ids: list[int], allowed_region_ids: set[int]
) -> list[dict]:
    """Ruxsat etilgan regionlar kesimida is_entered=False bo'lgan talabgorlar.

    Bir nechta smena bo'yicha ham ishlaydi — kun yoki sessiya bo'yicha
    aggregat holatda barcha smenalardagi kelmaganlarni qaytaradi.

    is_applied=True bo'lganlar ham testga kira olmaydigan kategoriya — ularni
    "kelmagan"lar ro'yxatiga kiritmaymiz (testga umuman kelmasligi
    rejalashtirilgan).

    Tartib: kun → smena_number → viloyat → bino → guruh → joy.
    """
    rows = db.execute(
        select(
            Student.id,
            Student.last_name,
            Student.first_name,
            Student.middle_name,
            Student.imei,
            Student.gr_n,
            Student.sp_n,
            Student.is_cheating,
            TestSessionSmena.day.label("test_day"),
            TestSessionSmena.number.label("smena_number"),
            Smena.name.label("smena_name"),
            Zone.name.label("zone_name"),
            Zone.number.label("zone_number"),
            Region.name.label("region_name"),
            Region.number.label("region_number"),
        )
        .join(TestSessionSmena, TestSessionSmena.id == Student.session_smena_id)
        .join(Smena, Smena.id == TestSessionSmena.test_smena_id)
        .join(Zone, Zone.id == Student.zone_id)
        .join(Region, Region.id == Zone.region_id)
        .where(
            Student.session_smena_id.in_(session_smena_ids),
            Region.id.in_(allowed_region_ids),
            Student.is_entered.is_(False),
            Student.is_applied.is_(False),
        )
        .order_by(
            TestSessionSmena.day,
            TestSessionSmena.number,
            Region.number,
            Region.name,
            Zone.number,
            Zone.name,
            Student.gr_n,
            Student.sp_n,
            Student.id,
        )
    ).all()

    result: list[dict] = []
    for r in rows:
        fio = " ".join(
            p for p in [r.last_name, r.first_name, r.middle_name] if p
        ).strip()
        result.append(
            {
                "fio": fio,
                "imei": r.imei or "",
                "gr_n": int(r.gr_n or 0),
                "sp_n": int(r.sp_n or 0),
                "region_name": r.region_name or "",
                "zone_name": r.zone_name or "",
                "test_day": r.test_day,
                "smena_number": int(r.smena_number or 0),
                "smena_name": r.smena_name or "",
                "is_cheating": bool(r.is_cheating),
            }
        )
    return result


def build_absentees_excel(
    db: Session,
    *,
    session_id: int,
    session_smena_id: int | None = None,
    test_day: date | None = None,
    allowed_region_ids: set[int],
) -> tuple[bytes, str, int]:
    """Kelmaganlar ro'yxatini .xlsx bytes shaklida qaytaradi.

    Tanlangan kontekst:
      - `session_smena_id` berilsa: bitta smena.
      - `test_day` berilsa: shu kunning barcha aktiv smenalari.
      - Ikkalasi `None` bo'lsa: sessiyaning barcha aktiv smenalari.

    Returns:
        (bytes, filename, absent_count)
    """
    if not allowed_region_ids:
        raise AbsenteesError("Foydalanuvchiga region biriktirilmagan")

    (
        test_session,
        smena_ids,
        resolved_day,
        smena_name,
        smena_number,
        scope,
    ) = _resolve_smenas(
        db,
        session_id=session_id,
        session_smena_id=session_smena_id,
        test_day=test_day,
    )
    test_name = (
        test_session.test.name if test_session.test else None
    ) or test_session.name

    rows = _fetch_absentees(
        db,
        session_smena_ids=smena_ids,
        allowed_region_ids=allowed_region_ids,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Kelmaganlar"

    # Aggregat (kun/sessiya) holatda Sana va Smena ustunlari ham qo'shiladi.
    is_aggr = scope != "smena"
    if is_aggr:
        columns = [
            ("№", 6),
            ("Sana", 14),
            ("Smena", 18),
            ("Viloyat", 24),
            ("Bino", 24),
            ("FIO", 60),
            ("IMEI (JShShIR)", 18),
            ("Gr", 8),
            ("Keldimi", 14),
            ("Chetlatish", 16),
        ]
    else:
        columns = [
            ("№", 6),
            ("Viloyat", 28),
            ("Bino", 28),
            ("FIO", 60),
            ("IMEI (JShShIR)", 18),
            ("Gr", 8),
            ("Keldimi", 14),
            ("Chetlatish", 16),
        ]
    n_cols = len(columns)
    last_col = get_column_letter(n_cols)

    # ── Sarlavha bloki: test nomi (1-qator) ─────────────────────────────
    if scope == "day":
        title_text = (
            f"KELMAGANLAR — {test_name} • "
            f"{resolved_day.isoformat() if resolved_day else '—'}"
        )
    elif scope == "total":
        title_text = f"KELMAGANLAR (UMUMIY) — {test_name}"
    else:
        title_text = f"KELMAGANLAR RO'YXATI — {test_name}"

    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[1].height = 30

    # 2-qator: sana
    ws.merge_cells(f"A2:{last_col}2")
    date_cell = ws["A2"]
    if scope == "total":
        date_cell.value = "Sana: Barcha kunlar"
    else:
        date_cell.value = (
            f"Test sanasi: {resolved_day.isoformat() if resolved_day else '—'}"
        )
    date_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    date_cell.fill = _META_FILL
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # 3-qator: smena
    ws.merge_cells(f"A3:{last_col}3")
    smena_cell = ws["A3"]
    if scope == "smena":
        smena_label = smena_name or "—"
        if smena_number:
            smena_label = f"#{smena_number} • {smena_label}"
        smena_cell.value = f"Smena: {smena_label}"
    elif scope == "day":
        smena_cell.value = "Smena: Shu sanadagi barcha smenalar"
    else:
        smena_cell.value = "Smena: Barcha smenalar"
    smena_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    smena_cell.fill = _META_FILL
    smena_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    # 4-qator: jami sanab ko'rsatish (kelmaganlar soni)
    ws.merge_cells(f"A4:{last_col}4")
    total_cell = ws["A4"]
    total_cell.value = f"Jami kelmaganlar: {len(rows)} ta"
    total_cell.font = Font(
        name="Calibri", size=11, bold=True, italic=True, color="1B5E20"
    )
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 20

    # 5-qator bo'sh oraliq
    ws.row_dimensions[5].height = 6

    # ── Jadval sarlavhalari (6-qator) ───────────────────────────────────
    header_row = 6
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER_HEADER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 26

    # ── Body ────────────────────────────────────────────────────────────
    data_start = header_row + 1
    cheat_font = Font(name="Calibri", size=11, bold=True, color="C62828")
    kelmadi_font = Font(name="Calibri", size=11, bold=True, color="EF6C00")
    # Aggregat va smena variantlari uchun ustun indekslari farq qiladi —
    # bitta joyda hisoblab olamiz.
    if is_aggr:
        col_imei = 7
        col_fio = 6
        col_keldi = 9
        col_chet = 10
        col_gr = 8
    else:
        col_imei = 5
        col_fio = 4
        col_keldi = 7
        col_chet = 8
        col_gr = 6

    for i, r in enumerate(rows, start=1):
        row_idx = data_start + i - 1
        cheating = bool(r.get("is_cheating"))
        if is_aggr:
            day_str = (
                r["test_day"].isoformat()
                if isinstance(r.get("test_day"), date)
                else _safe(r.get("test_day"))
            )
            sm_n = r.get("smena_number") or 0
            sm_label = r.get("smena_name") or ""
            smena_disp = (
                f"#{sm_n} • {sm_label}" if sm_n else sm_label
            )
            values = [
                i,
                day_str,
                smena_disp,
                _safe(r["region_name"]),
                _safe(r["zone_name"]),
                _safe(r["fio"]),
                _safe(r["imei"]),
                r["gr_n"] or "",
                "Kelmadi",
                "Chetlatilgan" if cheating else "",
            ]
        else:
            values = [
                i,
                _safe(r["region_name"]),
                _safe(r["zone_name"]),
                _safe(r["fio"]),
                _safe(r["imei"]),
                r["gr_n"] or "",
                "Kelmadi",
                "Chetlatilgan" if cheating else "",
            ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=11)
            cell.border = _BORDER_CELL
            if col_idx == col_imei:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == col_fio:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=False
                )
            elif col_idx == col_keldi:
                cell.font = kelmadi_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == col_chet:
                if cheating:
                    cell.font = cheat_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (1, col_gr):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            if i % 2 == 0:
                cell.fill = _ZEBRA_FILL
        ws.row_dimensions[row_idx].height = 20

    # Bo'sh holat — bitta xabar qatori
    if not rows:
        empty_row = data_start
        ws.merge_cells(
            start_row=empty_row, start_column=1, end_row=empty_row, end_column=n_cols
        )
        cell = ws.cell(
            row=empty_row, column=1, value="✓ Bu kontekstda kelmagan talabgor yo'q"
        )
        cell.font = Font(name="Calibri", size=12, italic=True, color="2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER_CELL
        ws.row_dimensions[empty_row].height = 28

    # Freeze panes — sarlavha + header doim ko'rinib tursin.
    ws.freeze_panes = ws.cell(row=data_start, column=1)
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # ── Fayl baytlari ───────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if scope == "day":
        safe_day = resolved_day.isoformat() if resolved_day else "no-date"
        filename = f"kelmaganlar_{safe_day}_umumiy.xlsx"
    elif scope == "total":
        filename = "kelmaganlar_umumiy.xlsx"
    else:
        safe_day = resolved_day.isoformat() if resolved_day else "no-date"
        filename = f"kelmaganlar_{safe_day}_smena_{smena_number or 'x'}.xlsx"
    return buffer.getvalue(), filename, len(rows)
