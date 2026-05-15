"""Excel orqali studentlarni yuklash (alternativ "manual" oqim).

Asosiy oqim:
1. .xlsx faylni openpyxl bilan parslaymiz.
2. Har bir qator validatsiya qilinadi: majburiy ustunlar, sana formati,
   regionda zona mavjudligi, sessiya ichida smena_number topiladimi.
3. Student va StudentPsData jadvallariga bulk-insert qilinadi.
4. Har bir student uchun GTSP API chaqiriladi (parallel: 5 ta thread).
   FIO va rasm yangilanadi; gender_id (key=1/2) o'rnatiladi.
5. Progress Redis'ga yoziladi — UI dagi mavjud progress widget shu kalitni
   o'qiydi.

Tashqi API (CEFR/MS/IIV) bilan ishlovchi `student_loader.py` mavjud — bu
modul shunga o'xshash shaklda, lekin manba sifatida foydalanuvchi yuborgan
Excel'ni oladi.
"""

from __future__ import annotations

import json
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any

import redis
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import settings
from app.models.gender import Gender
from app.models.region import Region
from app.models.smena import Smena
from app.models.student import Student
from app.models.student_ps_data import StudentPsData
from app.models.test_session import TestSession
from app.models.test_session_smena import TestSessionSmena
from app.models.zone import Zone
from app.services.gtsp_client import GtspError, GtspNotConfigured, fetch_gtsp_data

logger = logging.getLogger("faceid.services.excel_student_loader")

INSERT_BATCH_SIZE = 500
GTSP_WORKERS = 5
GTSP_TIMEOUT = 10.0

# Progress kaliti — student_loader.py BILAN BIR XIL bo'lishi shart, frontend
# `/student-load-progress` endpoint orqali shu kalitdan o'qiydi.
_PROGRESS_KEY = "student_load_progress:{session_id}"
_PROGRESS_TTL = 3600


class ExcelLoadError(Exception):
    def __init__(self, message: str) -> None:
        super().__init__(message)
        self.message = message


# ─── Excel parser ────────────────────────────────────────────────────


@dataclass
class ExcelRow:
    row_no: int  # Excel'dagi qator raqami (xato xabarlari uchun)
    last_name: str
    first_name: str
    middle_name: str | None
    imei: str | None
    ps_ser: str
    ps_num: str
    region_number: int
    zone_number: int | None
    smena_number: int
    gr_n: int
    e_date: date
    subject_name: str | None


# Header'larni normallashtirish — bo'sh joy, registr, alternativ yozuvlar.
_HEADER_ALIASES: dict[str, str] = {
    "last_name": "last_name",
    "lastname": "last_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "middle_name": "middle_name",
    "middlename": "middle_name",
    "imei": "imei",
    "ps_ser": "ps_ser",
    "psser": "ps_ser",
    "ps_num": "ps_num",
    "psnum": "ps_num",
    "region_number": "region_number",
    "region number": "region_number",
    "region": "region_number",
    "zone_number": "zone_number",
    "zone number": "zone_number",
    "zone": "zone_number",
    "smena_number": "smena_number",
    "smena number": "smena_number",
    "smena": "smena_number",
    "gr_n": "gr_n",
    "e_date": "e_date",
    "edate": "e_date",
    "subject_name": "subject_name",
    "subjectname": "subject_name",
    "subject": "subject_name",
}

# Majburiy maydonlar — yo'q bo'lsa qator o'tkazib yuboriladi va xato ro'yxatga
# qo'shiladi.
_REQUIRED_FIELDS = ("last_name", "first_name", "ps_ser", "ps_num", "smena_number", "e_date")


def _norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()


def _to_int(val: Any) -> int | None:
    if val is None or val == "":
        return None
    try:
        if isinstance(val, float):
            return int(val)
        return int(str(val).strip())
    except (TypeError, ValueError):
        return None


def _to_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _to_date(val: Any) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_excel(content: bytes) -> tuple[list[ExcelRow], list[str]]:
    """Excel byte'larini ro'yxat qatorlariga aylantirish.

    Returns:
        (rows, errors) — `errors` qatorlarni o'tkazib yuborish sabablarini
        ro'yxati. Validatsion xatoliklar bo'lsa ham qisman rows qaytaramiz
        — chaqiruvchi tomon qaror qiladi.
    """
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ExcelLoadError(f"Excel faylini o'qib bo'lmadi: {e}")

    ws = wb.active
    if ws is None:
        raise ExcelLoadError("Excel'da varaq topilmadi")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ExcelLoadError("Excel bo'sh — sarlavha qatori yo'q")

    # Normallashtirilgan ustun → index map
    col_index: dict[str, int] = {}
    for idx, raw in enumerate(header or []):
        key = _HEADER_ALIASES.get(_norm(raw))
        if key:
            col_index.setdefault(key, idx)

    missing = [f for f in _REQUIRED_FIELDS if f not in col_index]
    if missing:
        raise ExcelLoadError(
            "Excel sarlavhasida ustun(lar) topilmadi: " + ", ".join(missing)
        )

    def cell(row: tuple, key: str) -> Any:
        idx = col_index.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    parsed: list[ExcelRow] = []
    errors: list[str] = []

    for row_no, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None or c == "" for c in row):
            continue  # Bo'sh qator — o'tkazib yuboramiz

        last_name = _to_str(cell(row, "last_name")).upper()
        first_name = _to_str(cell(row, "first_name")).upper()
        middle_name = _to_str(cell(row, "middle_name")).upper() or None
        imei = _to_str(cell(row, "imei"))[:14] or None
        ps_ser = _to_str(cell(row, "ps_ser")).upper()
        ps_num = _to_str(cell(row, "ps_num"))
        region_number = _to_int(cell(row, "region_number"))
        zone_number = _to_int(cell(row, "zone_number"))
        smena_number = _to_int(cell(row, "smena_number"))
        gr_n = _to_int(cell(row, "gr_n")) or 0
        e_date = _to_date(cell(row, "e_date"))
        subject_name = _to_str(cell(row, "subject_name")) or None

        # Validatsiya
        problems: list[str] = []
        if not last_name:
            problems.append("last_name bo'sh")
        if not first_name:
            problems.append("first_name bo'sh")
        if not ps_ser:
            problems.append("ps_ser bo'sh")
        if not ps_num:
            problems.append("ps_num bo'sh")
        if smena_number is None:
            problems.append("smena number bo'sh yoki noto'g'ri")
        if e_date is None:
            problems.append("e_date noto'g'ri formatda")
        if region_number is None and "region_number" in col_index:
            # Region ixtiyoriy emas — agar ustun bor bo'lsa qiymat ham bo'lsin
            problems.append("region number bo'sh")

        if problems:
            errors.append(f"qator {row_no}: " + "; ".join(problems))
            continue

        parsed.append(
            ExcelRow(
                row_no=row_no,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
                imei=imei,
                ps_ser=ps_ser,
                ps_num=ps_num,
                region_number=int(region_number) if region_number is not None else 0,
                zone_number=zone_number,
                smena_number=int(smena_number),
                gr_n=gr_n,
                e_date=e_date,  # type: ignore[arg-type]
                subject_name=subject_name,
            )
        )

    return parsed, errors


# ─── Lookups ─────────────────────────────────────────────────────────


def _build_zone_lookup(db: Session) -> tuple[dict[tuple[int, int], int], dict[int, int]]:
    """((region.number, zone.number) → zone.id, region.number → first active zone.id) ikki map.

    Zone.number unique emas — shuning uchun (region, zone) tuple bilan kalitlaymiz.
    Agar Excel'da faqat region berilsa, fallback sifatida birinchi aktiv zonani olamiz.
    """
    rows = db.execute(
        select(Region.number, Zone.number, Zone.id, Zone.is_active)
        .join(Zone, Zone.region_id == Region.id)
        .order_by(Region.number, Zone.id)
    ).all()
    exact: dict[tuple[int, int], int] = {}
    first_in_region: dict[int, int] = {}
    for region_num, zone_num, zone_id, is_active in rows:
        if not is_active:
            continue
        exact.setdefault((int(region_num), int(zone_num or 0)), int(zone_id))
        first_in_region.setdefault(int(region_num), int(zone_id))
    return exact, first_in_region


def _build_session_smena_lookup(
    db: Session, session_id: int
) -> dict[tuple[int, date], int]:
    """((smena.number, day) → TestSessionSmena.id) shu sessiyaning smenalar uchun."""
    rows = (
        db.execute(
            select(TestSessionSmena)
            .join(Smena, Smena.id == TestSessionSmena.test_smena_id)
            .where(TestSessionSmena.test_session_id == session_id)
        )
        .scalars()
        .all()
    )
    out: dict[tuple[int, date], int] = {}
    for sm in rows:
        if sm.smena is None:
            continue
        out[(int(sm.smena.number), sm.day)] = int(sm.id)
    return out


def _build_gender_lookup(db: Session) -> dict[int, int]:
    """Gender.key → Gender.id (1=erkak, 2=ayol, boshqa=0)."""
    rows = db.execute(select(Gender.key, Gender.id)).all()
    return {int(k): int(i) for k, i in rows}


# ─── Redis progress ───────────────────────────────────────────────────


def _get_redis() -> redis.Redis | None:
    try:
        return redis.Redis.from_url(settings.REDIS_URL, decode_responses=True)
    except Exception:
        logger.warning("Redis ulanish xatosi")
        return None


def _set_progress(
    r: redis.Redis | None,
    session_id: int,
    *,
    current: int,
    total: int,
    status: str,
    message: str = "",
    skipped: int = 0,
) -> None:
    if r is None:
        return
    try:
        percent = round((current / total) * 100, 1) if total else 0
        percent = min(percent, 100.0)
        data = {
            "current": current,
            "total": total,
            "pages_done": 1 if status == "completed" else 0,
            "pages_total": 1,
            "skipped": skipped,
            "status": status,
            "message": message,
            "percent": percent,
        }
        r.set(
            _PROGRESS_KEY.format(session_id=session_id),
            json.dumps(data),
            ex=_PROGRESS_TTL,
        )
    except redis.ConnectionError:
        pass


# ─── Bulk insert ──────────────────────────────────────────────────────


def _insert_students(
    db: Session,
    session_id: int,
    rows: list[ExcelRow],
    zone_exact: dict[tuple[int, int], int],
    zone_first: dict[int, int],
    smena_map: dict[tuple[int, date], int],
) -> tuple[list[int], list[str]]:
    """Studentlarni va StudentPsData yozuvlarini batch'larda yaratish.

    Returns:
        (created_student_ids, errors)
    """
    errors: list[str] = []
    student_ids: list[int] = []

    student_rows: list[dict] = []
    ps_rows_pending: list[tuple[int, str, str]] = []  # (index_in_student_rows, ps_ser, ps_num)

    for row in rows:
        # Smena
        smena_key = (row.smena_number, row.e_date)
        session_smena_id = smena_map.get(smena_key)
        if session_smena_id is None:
            errors.append(
                f"qator {row.row_no}: smena #{row.smena_number} "
                f"({row.e_date}) sessiyada topilmadi"
            )
            continue

        # Zone
        zone_id: int | None = None
        if row.zone_number is not None:
            zone_id = zone_exact.get((row.region_number, row.zone_number))
        if zone_id is None:
            zone_id = zone_first.get(row.region_number)
        if zone_id is None:
            errors.append(
                f"qator {row.row_no}: region #{row.region_number} uchun zona topilmadi"
            )
            continue

        student_rows.append(
            {
                "session_smena_id": session_smena_id,
                "zone_id": zone_id,
                "last_name": row.last_name,
                "first_name": row.first_name,
                "middle_name": row.middle_name,
                "imei": row.imei,
                "gr_n": row.gr_n,
                "sp_n": 0,
                "s_code": 0,
                "e_date": row.e_date,
                "subject_id": 0,
                "subject_name": row.subject_name,
                "lang_id": 1,
                "level_id": 8,
                "is_ready": False,
                "is_face": False,
                "is_image": False,
                "is_cheating": False,
                "is_blacklist": False,
                "is_entered": False,
                "is_applied": False,
            }
        )
        ps_rows_pending.append((len(student_rows) - 1, row.ps_ser, row.ps_num))

    if not student_rows:
        return student_ids, errors

    # Batch insert — bulk_save_objects beradi qaytariladigan id larni,
    # bulk_insert_mappings tezroq lekin id qaytarmaydi. Bizga id kerak,
    # shuning uchun Student'larni alohida-alohida yaratamiz batch'larda.
    for batch_start in range(0, len(student_rows), INSERT_BATCH_SIZE):
        batch = student_rows[batch_start : batch_start + INSERT_BATCH_SIZE]
        objs = [Student(**data) for data in batch]
        db.add_all(objs)
        db.flush()  # id'larni olish uchun
        for idx, stu in enumerate(objs):
            real_idx = batch_start + idx
            student_ids.append(int(stu.id))
            ps_ser, ps_num = ps_rows_pending[real_idx][1], ps_rows_pending[real_idx][2]
            db.add(
                StudentPsData(
                    student_id=stu.id,
                    ps_ser=ps_ser[:5],
                    ps_num=ps_num[:10],
                )
            )
        db.commit()

    return student_ids, errors


# ─── GTSP enrichment ──────────────────────────────────────────────────


def _enrich_one(
    args: tuple[int, str | None, str, str],
) -> tuple[int, Any]:
    """Bitta studentni GTSP'dan boyitish (thread'da ishlaydi).

    Args tuple: (student_id, imei, ps_ser, ps_num).
    Returns: (student_id, GtspResult | Exception | None).
    """
    student_id, imei, ps_ser, ps_num = args
    ps_value = f"{ps_ser}{ps_num}"
    try:
        return student_id, fetch_gtsp_data(imei, ps_value, timeout=GTSP_TIMEOUT)
    except GtspNotConfigured:
        return student_id, None  # GTSP yo'q — sukut bilan o'tkazib yuboramiz
    except GtspError as e:
        return student_id, e
    except Exception as e:  # noqa: BLE001
        return student_id, e


def _apply_gtsp_result(
    db: Session,
    student_id: int,
    result: Any,
    gender_map: dict[int, int],
) -> bool:
    """Bitta student'ga GTSP natijasini DB'ga yozish. True = muvaffaqiyat."""
    if result is None or isinstance(result, Exception):
        return False
    student = db.get(Student, student_id)
    if student is None:
        return False
    ps_data = db.execute(
        select(StudentPsData).where(StudentPsData.student_id == student_id)
    ).scalar()
    if ps_data is None:
        return False

    if result.last_name:
        student.last_name = result.last_name
    if result.first_name:
        student.first_name = result.first_name
    if result.middle_name is not None:
        student.middle_name = result.middle_name
    if result.photo:
        ps_data.ps_img = result.photo
        student.is_image = True

    gender_key = result.sex if result.sex in (1, 2) else 0
    gid = gender_map.get(gender_key)
    if gid:
        ps_data.gender_id = gid
    return True


def _enrich_via_gtsp(
    db: Session,
    session_id: int,
    student_ids: list[int],
    r: redis.Redis | None,
) -> dict[str, int]:
    """Barcha yangi studentlarni GTSP orqali boyitish.

    Parallel: GTSP_WORKERS thread. DB yozish — sequential (main thread).
    """
    if not student_ids:
        return {"enriched": 0, "failed": 0}

    # Bir martalik metadata yig'amiz
    rows = db.execute(
        select(Student.id, Student.imei, StudentPsData.ps_ser, StudentPsData.ps_num)
        .join(StudentPsData, StudentPsData.student_id == Student.id)
        .where(Student.id.in_(student_ids))
    ).all()
    work_items: list[tuple[int, str | None, str, str]] = [
        (int(sid), imei, ps_ser, ps_num) for sid, imei, ps_ser, ps_num in rows
    ]

    gender_map = _build_gender_lookup(db)
    total = len(work_items)
    enriched = 0
    failed = 0

    with ThreadPoolExecutor(max_workers=GTSP_WORKERS) as ex:
        future_map = {ex.submit(_enrich_one, item): item[0] for item in work_items}
        for i, fut in enumerate(as_completed(future_map), start=1):
            student_id, result = fut.result()
            ok = _apply_gtsp_result(db, student_id, result, gender_map)
            if ok:
                enriched += 1
            else:
                failed += 1
            # Har 20 ta da commit + progress yangilaymiz
            if i % 20 == 0 or i == total:
                try:
                    db.commit()
                except Exception:
                    db.rollback()
                    logger.exception("GTSP enrich commit xatosi")
                _set_progress(
                    r,
                    session_id,
                    current=i,
                    total=total,
                    status="processing",
                    message=f"GTSP: {i}/{total} ({enriched} muvaffaqiyat, {failed} xato)",
                )

    return {"enriched": enriched, "failed": failed}


# ─── Asosiy entry point ───────────────────────────────────────────────


def load_students_from_excel(
    db: Session,
    session: TestSession,
    excel_content: bytes,
) -> dict[str, Any]:
    """Excel'dan studentlarni yuklash + GTSP boyitish + progress.

    Returns:
        {"inserted": int, "enriched": int, "failed": int, "errors": [...]}

    Raises:
        ExcelLoadError: parse xatosi yoki yuklab bo'lmaydigan holat.
    """
    r = _get_redis()
    _set_progress(
        r, session.id, current=0, total=0, status="processing",
        message="Excel parslangmoqda...",
    )

    rows, parse_errors = parse_excel(excel_content)
    if not rows:
        _set_progress(
            r,
            session.id,
            current=0,
            total=0,
            status="error",
            message="Excelda yaroqli qator topilmadi",
        )
        raise ExcelLoadError(
            "Yaroqli qatorlar topilmadi: " + ("; ".join(parse_errors)[:300] or "bo'sh fayl")
        )

    logger.info(
        "Excel #%d: %d ta qator parse qilindi, %d ta parse xatosi",
        session.id,
        len(rows),
        len(parse_errors),
    )

    zone_exact, zone_first = _build_zone_lookup(db)
    smena_map = _build_session_smena_lookup(db, session.id)
    if not smena_map:
        _set_progress(
            r,
            session.id,
            current=0,
            total=len(rows),
            status="error",
            message="Sessiyada smena topilmadi",
        )
        raise ExcelLoadError(
            "Bu sessiyada smena qo'shilmagan — avval smena qo'shing"
        )

    _set_progress(
        r,
        session.id,
        current=0,
        total=len(rows),
        status="processing",
        message="Studentlar bazaga yozilmoqda...",
    )

    inserted_ids, insert_errors = _insert_students(
        db, session.id, rows, zone_exact, zone_first, smena_map
    )
    inserted = len(inserted_ids)
    logger.info(
        "Excel #%d: %d ta student yaratildi, %d ta insert xatosi",
        session.id,
        inserted,
        len(insert_errors),
    )

    _set_progress(
        r,
        session.id,
        current=0,
        total=inserted,
        status="processing",
        message=f"{inserted} ta student yaratildi — GTSP ma'lumotlari yuklanmoqda",
        skipped=len(parse_errors) + len(insert_errors),
    )

    enrich_stats = _enrich_via_gtsp(db, session.id, inserted_ids, r)

    _set_progress(
        r,
        session.id,
        current=inserted,
        total=inserted,
        status="completed",
        message=(
            f"Yuklandi: {inserted} ta; GTSP muvaffaqiyatli: "
            f"{enrich_stats['enriched']}, xato: {enrich_stats['failed']}"
        ),
        skipped=len(parse_errors) + len(insert_errors),
    )

    return {
        "inserted": inserted,
        "enriched": enrich_stats["enriched"],
        "failed": enrich_stats["failed"],
        "errors": parse_errors + insert_errors,
    }
