"""Talabalarning passport (ps_ser / ps_num) ma'lumotlarini ommaviy yangilash.

Foydalanish konteksti: ba'zi talabalarning passport seriyasi/raqami
eskirgan bo'ladi. Operator Excel shablon (`jshshir, ps_ser, ps_num`),
Excel'dan nusxalab qo'yilgan (paste) qatorlar yoki faqat PINFL ro'yxatini
yuboradi — oxirgi holatda seriya/raqam e-gov PSN API'dan olinadi
(`update_session_passports_from_psn`). Har uchala yo'lda ham `jshshir`
(= `Student.imei`) bo'yicha shu sessiyadagi talaba topilib, uning
`StudentPsData` yozuvidagi `ps_ser`/`ps_num` yangilanadi.

Bu jarayon yengil (faqat DB UPDATE, yuz/embedding qayta ishlash yo'q),
shuning uchun sinxron bajariladi — Celery talab qilinmaydi.
"""

from __future__ import annotations

import base64
import binascii
import logging
import re
from io import BytesIO

from sqlalchemy import select as sa_select
from sqlalchemy.orm import Session

from app.models.student import Student
from app.models.student_ps_data import StudentPsData
from app.models.test_session_smena import TestSessionSmena
from app.services.egov_psn_client import birth_date_from_pinfl, fetch_documents

logger = logging.getLogger(__name__)

# Maydon uzunligi cheklovlari — DB ustun o'lchamlariga mos
# (Student.imei = String(14), StudentPsData.ps_ser = String(5), ps_num = String(10)).
_MAX_JSHSHIR = 14
_MAX_PS_SER = 5
_MAX_PS_NUM = 10

# Excel sarlavhalari uchun alias'lar — turli shablonlarni ham qabul qilamiz.
_HEADER_ALIASES: dict[str, str] = {
    "jshshir": "jshshir",
    "jshshr": "jshshir",
    "pinfl": "jshshir",
    "imei": "jshshir",
    "ps_ser": "ps_ser",
    "psser": "ps_ser",
    "seria": "ps_ser",
    "seriya": "ps_ser",
    "series": "ps_ser",
    "pasport_seriyasi": "ps_ser",
    "ps_num": "ps_num",
    "psnum": "ps_num",
    "raqam": "ps_num",
    "number": "ps_num",
    "pasport_raqami": "ps_num",
}


def _norm_header(value: object) -> str:
    """Sarlavha matnini kalitga aylantiradi (kichik harf, bo'shliq/nuqta → '_')."""
    text = str(value or "").strip().lower()
    for ch in (" ", ".", "-", "/"):
        text = text.replace(ch, "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")


# "Seriya+raqam" bitta katakka birlashgan holat: AA1234567 / AA 1234567 / AA-1234567.
# Harflar (seriya) + ixtiyoriy bo'shliq/tire + raqamlar (raqam).
_COMBINED_PASSPORT_RE = re.compile(r"^([A-Za-z]{1,5})[\s\-]*(\d{1,10})$")


def split_combined_passport(ps_ser: str, ps_num: str) -> tuple[str, str]:
    """`ps_num` bo'sh va `ps_ser` ichida seriya+raqam birga kelgan bo'lsa, ajratadi.

    Misol: ("AA1234567", "") -> ("AA", "1234567").
    `ps_num` allaqachon to'ldirilgan bo'lsa — hech narsa o'zgartirmaydi (to'g'ri
    qatorlarga tegmaymiz). Naqshga mos kelmasa ham o'zgarishsiz qaytaradi.
    """
    if ps_num or not ps_ser:
        return ps_ser, ps_num
    m = _COMBINED_PASSPORT_RE.match(ps_ser.strip())
    if not m:
        return ps_ser, ps_num
    return m.group(1).upper(), m.group(2)


def clean_cell(value: object) -> str:
    """Yacheykadagi qiymatni tozalaydi. Excel sonlarini ('123.0') matnga keltiradi."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def validate_jshshir(jshshir: str) -> str | None:
    """JSHSHIR ni tekshiradi. Xato bo'lsa — sabab matnini, aks holda None."""
    if not jshshir:
        return "JSHSHIR bo'sh"
    if len(jshshir) > _MAX_JSHSHIR:
        return f"JSHSHIR {_MAX_JSHSHIR} belgidan oshmasin"
    if not jshshir.isdigit():
        return "JSHSHIR faqat raqamlardan iborat bo'lishi kerak"
    return None


def validate_row(jshshir: str, ps_ser: str, ps_num: str) -> str | None:
    """Bitta qatorni tekshiradi. Xato bo'lsa — sabab matnini, aks holda None qaytaradi."""
    err = validate_jshshir(jshshir)
    if err:
        return err
    if not ps_ser:
        return "Pasport seriyasi bo'sh"
    if len(ps_ser) > _MAX_PS_SER:
        return f"Pasport seriyasi {_MAX_PS_SER} belgidan oshmasin"
    if not ps_num:
        return "Pasport raqami bo'sh"
    if len(ps_num) > _MAX_PS_NUM:
        return f"Pasport raqami {_MAX_PS_NUM} belgidan oshmasin"
    return None


def parse_passport_excel(content: bytes) -> tuple[list[dict], list[str]]:
    """`.xlsx` baytlaridan `[{jshshir, ps_ser, ps_num}, ...]` qatorlarini ajratadi.

    Qaytaradi: (rows, errors). `errors` — fayl darajasidagi (sarlavha topilmadi va h.k.)
    muammolar; qator darajasidagi validatsiya `update_session_passports` da bo'ladi.
    """
    from openpyxl import load_workbook

    errors: list[str] = []
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Passport Excel o'qishda xato: %s", exc)
        return [], ["Excel faylni o'qib bo'lmadi — fayl buzilgan yoki noto'g'ri formatda"]

    ws = wb.active
    if ws is None:
        return [], ["Excel'da varaq topilmadi"]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], ["Excel bo'sh"]

    col_index: dict[str, int] = {}
    for idx, raw in enumerate(header):
        key = _HEADER_ALIASES.get(_norm_header(raw))
        if key and key not in col_index:
            col_index[key] = idx

    missing = [c for c in ("jshshir", "ps_ser", "ps_num") if c not in col_index]
    if missing:
        return [], [
            "Excel'da ustunlar topilmadi: "
            + ", ".join(missing)
            + ". Kerakli sarlavhalar: jshshir, ps_ser, ps_num"
        ]

    rows: list[dict] = []
    for raw in rows_iter:
        if raw is None:
            continue

        def cell(key: str) -> str:
            i = col_index[key]
            return clean_cell(raw[i]) if i < len(raw) else ""

        jshshir = cell("jshshir")
        ps_ser = cell("ps_ser").upper()
        ps_num = cell("ps_num")
        if not (jshshir or ps_ser or ps_num):
            continue  # butunlay bo'sh qator — o'tkazib yuboramiz
        rows.append({"jshshir": jshshir, "ps_ser": ps_ser, "ps_num": ps_num})

    wb.close()
    return rows, errors


def update_session_passports(
    db: Session, session_id: int, rows: list[dict]
) -> dict:
    """`rows` ichidagi passport ma'lumotlarini sessiya talabalariga qo'llaydi.

    Moslashtirish: `Student.imei == jshshir` VA talaba shu `session_id` ga tegishli.
    Topilgan har bir talabaning `StudentPsData` yozuvi yangilanadi (yo'q bo'lsa
    yaratiladi).

    Qaytaradi:
        {
          "total": int,            # yuborilgan qatorlar soni
          "updated": int,          # yangilangan talaba yozuvlari soni
          "not_found": [jshshir],  # sessiyada topilmagan JSHSHIR'lar
          "invalid": [{"row", "jshshir", "error"}],  # validatsiyadan o'tmaganlar
        }
    """
    total = len(rows)
    invalid: list[dict] = []
    not_found: list[str] = []

    # 1) Qatorlarni tozalab/validatsiya qilamiz; yaroqlilarini jshshir bo'yicha yig'amiz.
    #    Bir xil jshshir takror kelsa — oxirgisi ustun bo'ladi.
    valid: dict[str, dict] = {}
    for idx, row in enumerate(rows, start=1):
        jshshir = clean_cell(row.get("jshshir"))
        ps_ser = clean_cell(row.get("ps_ser")).upper()
        ps_num = clean_cell(row.get("ps_num"))
        # Seriya+raqam bitta katakka birlashib kelgan bo'lsa — ajratamiz.
        ps_ser, ps_num = split_combined_passport(ps_ser, ps_num)
        err = validate_row(jshshir, ps_ser, ps_num)
        if err:
            invalid.append({"row": idx, "jshshir": jshshir, "error": err})
            continue
        valid[jshshir] = {"ps_ser": ps_ser, "ps_num": ps_num}

    if not valid:
        return {"total": total, "updated": 0, "not_found": [], "invalid": invalid}

    # 2) Shu sessiyaga tegishli, jshshir'i mos keladigan talabalarni topamiz.
    smena_subq = sa_select(TestSessionSmena.id).where(
        TestSessionSmena.test_session_id == session_id
    )
    students = (
        db.scalars(
            sa_select(Student).where(
                Student.session_smena_id.in_(smena_subq),
                Student.imei.in_(list(valid.keys())),
            )
        )
        .all()
    )

    by_imei: dict[str, list[Student]] = {}
    for st in students:
        by_imei.setdefault(st.imei or "", []).append(st)

    # 3) Mavjud ps_data yozuvlarini bitta so'rovda olib kelamiz.
    student_ids = [st.id for st in students]
    ps_by_student: dict[int, StudentPsData] = {}
    if student_ids:
        for ps in db.scalars(
            sa_select(StudentPsData).where(StudentPsData.student_id.in_(student_ids))
        ).all():
            ps_by_student[ps.student_id] = ps

    # 4) Yangilash.
    updated = 0
    for jshshir, data in valid.items():
        matched = by_imei.get(jshshir)
        if not matched:
            not_found.append(jshshir)
            continue
        for st in matched:
            ps = ps_by_student.get(st.id)
            if ps is None:
                ps = StudentPsData(
                    student_id=st.id, ps_ser=data["ps_ser"], ps_num=data["ps_num"]
                )
                db.add(ps)
                ps_by_student[st.id] = ps
            else:
                ps.ps_ser = data["ps_ser"]
                ps.ps_num = data["ps_num"]
            updated += 1

    db.commit()
    logger.info(
        "Passport yangilash: session=%d, total=%d, updated=%d, not_found=%d, invalid=%d",
        session_id, total, updated, len(not_found), len(invalid),
    )
    return {
        "total": total,
        "updated": updated,
        "not_found": not_found,
        "invalid": invalid,
    }


# ─── PSN (e-gov) orqali passport yangilash ───────────────────────────────

# PINFL aynan 14 xonali bo'lishi shart — PSN so'rovi uchun tug'ilgan sana
# shu raqamdan hisoblanadi.
_PINFL_LEN = 14


def validate_pinfl(pinfl: str) -> str | None:
    """PSN so'rovi uchun PINFL ni tekshiradi. Xato bo'lsa — sabab matni."""
    if not pinfl:
        return "PINFL bo'sh"
    if not pinfl.isdigit():
        return "PINFL faqat raqamlardan iborat bo'lishi kerak"
    if len(pinfl) != _PINFL_LEN:
        return f"PINFL {_PINFL_LEN} ta raqamdan iborat bo'lishi kerak"
    if birth_date_from_pinfl(pinfl) is None:
        return "PINFL dan tug'ilgan sana aniqlanmadi"
    return None


def _session_student_imeis(
    db: Session, session_id: int, candidates: list[str]
) -> set[str]:
    """`candidates` ichidan shu sessiyada mavjud bo'lgan IMEI (PINFL) larni qaytaradi."""
    if not candidates:
        return set()
    smena_subq = sa_select(TestSessionSmena.id).where(
        TestSessionSmena.test_session_id == session_id
    )
    rows = db.scalars(
        sa_select(Student.imei).where(
            Student.session_smena_id.in_(smena_subq),
            Student.imei.in_(candidates),
        )
    ).all()
    return {imei for imei in rows if imei}


def update_session_passports_from_psn(
    db: Session, session_id: int, pinfls: list[str]
) -> dict:
    """PINFL ro'yxati bo'yicha passportni e-gov PSN dan olib, sessiyaga qo'llaydi.

    Bosqichlar:
      1. PINFL larni tozalash/validatsiya, takrorlarni olib tashlash.
      2. Sessiyada bor-yo'qligini OLDINDAN tekshirish — sessiyada yo'q talaba
         uchun tashqi API ga so'rov yubormaymiz (kvota tejaladi, javob tezroq).
      3. Qolganlari uchun PSN dan `ps_ser`/`ps_num` olish (parallel).
      4. Olingan qatorlarni `update_session_passports` orqali yozish.

    Qaytaradi: `update_session_passports` summary'si + `failed` —
    PSN javob bermagan yoki rad etgan PINFL lar `{"pinfl", "error"}` ko'rinishida.
    """
    total = len(pinfls)
    invalid: list[dict] = []
    failed: list[dict] = []

    # 1) Tozalash + validatsiya. Takror PINFL bir marta so'raladi.
    candidates: list[str] = []
    seen: set[str] = set()
    for idx, raw in enumerate(pinfls, start=1):
        pinfl = clean_cell(raw)
        err = validate_pinfl(pinfl)
        if err:
            invalid.append({"row": idx, "jshshir": pinfl, "error": err})
            continue
        if pinfl in seen:
            continue
        seen.add(pinfl)
        candidates.append(pinfl)

    if not candidates:
        return {
            "total": total,
            "updated": 0,
            "not_found": [],
            "invalid": invalid,
            "failed": failed,
        }

    # 2) Sessiyada yo'qlarini ajratib olamiz — ular uchun PSN chaqirilmaydi.
    present = _session_student_imeis(db, session_id, candidates)
    targets = [p for p in candidates if p in present]
    not_found = [p for p in candidates if p not in present]

    if not targets:
        return {
            "total": total,
            "updated": 0,
            "not_found": not_found,
            "invalid": invalid,
            "failed": failed,
        }

    # 3) PSN dan olish (xatolar PINFL kesimida qaytadi — batch to'xtamaydi).
    documents, errors = fetch_documents(targets)
    failed.extend(
        {"pinfl": pinfl, "error": reason} for pinfl, reason in errors.items()
    )

    # 4) Yozishdan oldin oxirgi validatsiya — DB ustun o'lchamlariga mos kelsin.
    rows: list[dict] = []
    for pinfl, doc in documents.items():
        err = validate_row(pinfl, doc.ps_ser, doc.ps_num)
        if err:
            failed.append({"pinfl": pinfl, "error": err})
            continue
        rows.append(
            {"jshshir": pinfl, "ps_ser": doc.ps_ser, "ps_num": doc.ps_num}
        )

    if not rows:
        return {
            "total": total,
            "updated": 0,
            "not_found": not_found,
            "invalid": invalid,
            "failed": failed,
        }

    summary = update_session_passports(db, session_id, rows)

    # 2-bosqichdan keyin talaba o'chirilgan bo'lsa — bu yerda ham chiqishi mumkin.
    not_found.extend(summary["not_found"])
    # Ichki validatsiya xatolari (kutilmaydi, chunki yuqorida tekshirdik) —
    # foydalanuvchi kiritgan qator emas, PSN ma'lumoti, shuning uchun `failed` ga.
    failed.extend(
        {"pinfl": item["jshshir"], "error": item["error"]}
        for item in summary["invalid"]
    )

    logger.info(
        "PSN passport yangilash: session=%d, total=%d, so'ralgan=%d, "
        "updated=%d, not_found=%d, invalid=%d, failed=%d",
        session_id, total, len(targets), summary["updated"],
        len(not_found), len(invalid), len(failed),
    )
    return {
        "total": total,
        "updated": summary["updated"],
        "not_found": not_found,
        "invalid": invalid,
        "failed": failed,
    }


# ─── Passport RASMI (ps_img) ommaviy yangilash ────────────────────────────

# Bitta rasm uchun dekodlangandan keyingi hajm chegaralari.
_MIN_IMAGE_BYTES = 512           # 512 B dan kichigi rasm bo'lishi mumkin emas
_MAX_IMAGE_BYTES = 3 * 1024 * 1024  # 3 MB

# Ko'p ishlatiladigan rasm formatlarining "sehrli baytlari" (magic bytes).
# Base64 noto'g'ri (masalan matn yoki PDF) bo'lsa — DB ga yozmaymiz, chunki
# keyinchalik embedding bosqichida dekodlash xatosi bo'ladi.
_IMAGE_SIGNATURES: tuple[tuple[bytes, str], ...] = (
    (b"\xff\xd8\xff", "JPEG"),
    (b"\x89PNG\r\n\x1a\n", "PNG"),
    (b"BM", "BMP"),
    (b"GIF87a", "GIF"),
    (b"GIF89a", "GIF"),
)

# `data:image/jpeg;base64,....` shaklidagi data-URL prefiksi.
_DATA_URL_RE = re.compile(r"^data:[^;,]*;base64,", re.IGNORECASE)


def _detect_image_format(raw: bytes) -> str | None:
    """Bayt oqimidan rasm formatini aniqlaydi. Rasm bo'lmasa — None."""
    for sig, name in _IMAGE_SIGNATURES:
        if raw.startswith(sig):
            return name
    # WEBP: "RIFF" + 4 bayt hajm + "WEBP"
    if raw[:4] == b"RIFF" and raw[8:12] == b"WEBP":
        return "WEBP"
    return None


def decode_passport_image(value: str) -> tuple[bytes | None, str | None]:
    """base64 satrni (data-URL yoki sof base64) rasm baytlariga aylantiradi.

    Qaytaradi: `(raw_bytes, None)` yoki xato bo'lsa `(None, "sabab")`.
    """
    text = (value or "").strip()
    if not text:
        return None, "Rasm bo'sh"

    text = _DATA_URL_RE.sub("", text)
    # Excel/matn muharrirlaridan nusxalanganda base64 ichiga qator uzilishi va
    # bo'shliqlar tushib qolishi mumkin — ularni olib tashlaymiz.
    text = re.sub(r"\s+", "", text)
    # URL-safe base64 (`-` va `_`) ni ham qabul qilamiz.
    text = text.replace("-", "+").replace("_", "/")
    # To'ldiruvchi `=` yetishmasa — qo'shib qo'yamiz.
    if len(text) % 4:
        text += "=" * (4 - len(text) % 4)

    try:
        raw = base64.b64decode(text, validate=True)
    except (binascii.Error, ValueError):
        return None, "Base64 formati noto'g'ri"

    if len(raw) < _MIN_IMAGE_BYTES:
        return None, "Rasm juda kichik (buzilgan bo'lishi mumkin)"
    if len(raw) > _MAX_IMAGE_BYTES:
        return None, f"Rasm hajmi {_MAX_IMAGE_BYTES // (1024 * 1024)}MB dan oshmasin"
    if _detect_image_format(raw) is None:
        return None, "Bu rasm emas (JPEG/PNG/WEBP/BMP/GIF kutilgan)"
    return raw, None


def update_session_passport_images(
    db: Session, session_id: int, rows: list[dict]
) -> dict:
    """`rows` ichidagi base64 rasmlarni sessiya talabalarining `ps_img` iga yozadi.

    Har bir qator: `{"jshshir": str, "image": str}`. Rasm `StudentPsData.ps_img`
    ga odatdagidek BLOB (xom bayt) ko'rinishida saqlanadi — base64 emas.

    Rasm almashgani uchun eski `embedding` yaroqsiz bo'lib qoladi: uni tozalab,
    talabani `is_ready=False` qilamiz — shunda embedding bosqichini qayta ishga
    tushirganda ("faqat tayyor bo'lmaganlar") bu talabalar qayta hisoblanadi.

    Qaytaradi: `update_session_passports` bilan bir xil summary.
    """
    total = len(rows)
    invalid: list[dict] = []
    not_found: list[str] = []

    # 1) Validatsiya + dekodlash. Bir xil jshshir takror kelsa — oxirgisi ustun.
    valid: dict[str, bytes] = {}
    for idx, row in enumerate(rows, start=1):
        jshshir = clean_cell(row.get("jshshir"))
        err = validate_jshshir(jshshir)
        if err:
            invalid.append({"row": idx, "jshshir": jshshir, "error": err})
            continue
        raw, img_err = decode_passport_image(str(row.get("image") or ""))
        if img_err or raw is None:
            invalid.append({"row": idx, "jshshir": jshshir, "error": img_err or "Rasm yaroqsiz"})
            continue
        valid[jshshir] = raw

    if not valid:
        return {"total": total, "updated": 0, "not_found": [], "invalid": invalid}

    # 2) Shu sessiyaga tegishli, jshshir'i mos keladigan talabalar.
    smena_subq = sa_select(TestSessionSmena.id).where(
        TestSessionSmena.test_session_id == session_id
    )
    students = db.scalars(
        sa_select(Student).where(
            Student.session_smena_id.in_(smena_subq),
            Student.imei.in_(list(valid.keys())),
        )
    ).all()

    by_imei: dict[str, list[Student]] = {}
    for st in students:
        by_imei.setdefault(st.imei or "", []).append(st)

    # 3) Mavjud ps_data yozuvlarini bitta so'rovda olib kelamiz.
    student_ids = [st.id for st in students]
    ps_by_student: dict[int, StudentPsData] = {}
    if student_ids:
        for ps in db.scalars(
            sa_select(StudentPsData).where(StudentPsData.student_id.in_(student_ids))
        ).all():
            ps_by_student[ps.student_id] = ps

    # 4) Yangilash.
    updated = 0
    for jshshir, raw in valid.items():
        matched = by_imei.get(jshshir)
        if not matched:
            not_found.append(jshshir)
            continue
        for st in matched:
            ps = ps_by_student.get(st.id)
            if ps is None:
                # ps_ser/ps_num NOT NULL — rasm yo'li orqali kelganda ular
                # noma'lum, shuning uchun bo'sh qiymat bilan yozuv ochamiz.
                ps = StudentPsData(student_id=st.id, ps_ser="", ps_num="")
                db.add(ps)
                ps_by_student[st.id] = ps
            ps.ps_img = raw
            ps.embedding = None  # eski embedding endi boshqa rasmga tegishli
            st.is_image = True
            st.is_face = False
            st.is_ready = False
            updated += 1

    db.commit()
    logger.info(
        "Passport rasm yangilash: session=%d, total=%d, updated=%d, not_found=%d, invalid=%d",
        session_id, total, updated, len(not_found), len(invalid),
    )
    return {
        "total": total,
        "updated": updated,
        "not_found": not_found,
        "invalid": invalid,
    }
