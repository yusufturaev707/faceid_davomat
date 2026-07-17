"""Natija uchun tahlil natijasini .xlsx hisobotiga eksport qilish."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.result_analysis import AnalysisMode, ResultAnalysisResponse

MODE_LABELS = {
    AnalysisMode.IN_FACE_NOT_EXCLUDED_NO_RESULT: (
        "Faceda bor - chetlatilmagan - natija chiqmagan"
    ),
    AnalysisMode.IN_FACE_EXCLUDED_HAS_RESULT: (
        "Faceda bor - chetlatilgan - natija chiqqan"
    ),
    AnalysisMode.NOT_IN_FACE_HAS_RESULT: "Faceda yo'q - natija chiqqan",
}

_HEADERS = [
    ("№", 6),
    ("Familiya", 18),
    ("Ism", 16),
    ("Sharif", 18),
    ("IMEI", 18),
    ("Viloyat", 20),
    ("Bino", 22),
    ("Test kuni", 13),
    ("Smena", 12),
    ("abitur_id", 14),
    ("tday", 13),
    ("deleted", 10),
    ("img (rasm)", 40),
]


def _img_url(base: str, img: str | None) -> str:
    """BASE_IMG_URL va img qiymatini bitta URL ga birlashtiradi."""
    if not base or not img:
        return ""
    return f"{base.rstrip('/')}/{img.lstrip('/')}"

_TITLE_FILL = PatternFill("solid", fgColor="1D4ED8")
_META_FILL = PatternFill("solid", fgColor="2563EB")
_HEADER_FILL = PatternFill("solid", fgColor="2563EB")
_ZEBRA_FILL = PatternFill("solid", fgColor="EFF6FF")
_THIN = Side(style="thin", color="BFDBFE")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")


def _fmt_day(value: str | None) -> str:
    """ISO "YYYY-MM-DD" -> "DD.MM.YYYY"; boshqasi o'zgarmaydi."""
    if not value:
        return ""
    v = value.strip()
    if len(v) >= 10 and v[4] == "-" and v[7] == "-":
        return f"{v[8:10]}.{v[5:7]}.{v[0:4]}"
    return v


def build_result_analysis_excel(
    resp: ResultAnalysisResponse, *, base_img_url: str = ""
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tahlil"
    ncols = len(_HEADERS)

    # 1-qator: sarlavha
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, "Natija uchun tahlil")
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = _TITLE_FILL
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 26

    # 2-qator: tahlil turi + statistika
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    meta = (
        f"{MODE_LABELS.get(resp.mode, resp.mode)}   |   "
        f"Ko'lamdagi talabalar: {resp.scope_total}   |   "
        f"Joylangan imei: {resp.pasted_total}   |   "
        f"Natija chiqqan: {resp.pasted_result_count}   |   "
        f"Topildi: {resp.count}"
    )
    c = ws.cell(2, 1, meta)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = _META_FILL
    c.alignment = _LEFT
    ws.row_dimensions[2].height = 20

    # 3-qator: ustun sarlavhalari
    header_row = 3
    for i, (title, width) in enumerate(_HEADERS, start=1):
        cell = ws.cell(header_row, i, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(i)].width = width

    # Ma'lumot qatorlari
    for idx, it in enumerate(resp.items, start=1):
        row = header_row + idx
        img_url = _img_url(base_img_url, it.img)
        values = [
            idx,
            it.last_name or "",
            it.first_name or "",
            it.middle_name or "",
            it.imei or "",
            it.region_name or "",
            it.zone_name or "",
            _fmt_day(it.test_day),
            it.smena_name or "",
            it.abitur_id or "",
            _fmt_day(it.tday),
            it.deleted or "",
            img_url,
        ]
        for i, v in enumerate(values, start=1):
            cell = ws.cell(row, i, v)
            cell.border = _BORDER
            cell.alignment = _CENTER if i in (1, 8, 9, 11, 12) else _LEFT
            if idx % 2 == 0:
                cell.fill = _ZEBRA_FILL
        # img ustuni — bosiladigan URL havola.
        if img_url:
            img_cell = ws.cell(row, ncols)
            img_cell.hyperlink = img_url
            img_cell.font = Font(color="1D4ED8", underline="single")

    ws.freeze_panes = "A4"

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
