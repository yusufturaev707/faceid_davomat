"""Sessiya statistika dashboardini Excel (.xlsx) hisobotiga yozish.

`C:\\FastApi\\faceid\\1-смена_бошланиши (8-10 гача).xlsx` shabloniga mos rasmiy
"МАЪЛУМОТ" hujjati ko'rinishida yig'iladi:

  1-qator   — test sessiya nomi
  2-qator   — "МАЪЛУМОТ"
  3-qator   — o'ng burchakda sana/vaqt/smena ("2026 йил 10 июль, соат 11:10, 1-смена")
  4-6-qator — jadval sarlavhalari (ko'p qatorli, merge bilan)
  7+        — har bir viloyat bo'yicha statistika
  oxirgi    — "Жами:" yig'indi qatori

Ustunlar semantikasi (shablon formulalariga mos):
  C  Иштирок этадиган абитуриентлар сони  = total
  D  Иштирок этганлар (сони)              = C - F  (formula)        ← haqiqatda testdan o'tganlar
  E  Иштирок этганлар (%)                 = D*100/C
  F  Иштирок этмаганлар (сони)            = not_attended + chetlatilgan
  G  Иштирок этмаганлар (%)               = F*100/C
  H  Тест синовларида қатнашмаганлар сони = C - D - I - J  (formula)
  I  Четлатилганлар — ҳудудга киришда     = cheating.at_entry
  J  Четлатилганлар — тест жараёнида      = cheating.during_test

`dashboard_stats.attended` chetlatilganlarni ham o'z ichiga oladi, shuning uchun
"haqiqatda testdan o'tganlar" (D) = attended - cheating, "иштирок этмаганлар"
(F) = not_attended + cheating — shablon (F = H + I + J) bilan to'liq mos keladi.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.dashboard_stats import DashboardStatsResponse

# O'zbekiston vaqti (UZT) — UTC+5, DST yo'q (app.crud.student bilan bir xil)
UZ_TZ = timezone(timedelta(hours=5))

# Cyrillic (rus uslubidagi) oy nomlari — shablon "июль" yozuviga mos
_MONTHS_RU = (
    "", "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
)

# O'zbek lotin oy nomlari (lotin alifbosida hisobot uchun)
_MONTHS_LATIN = (
    "", "yanvar", "fevral", "mart", "aprel", "may", "iyun",
    "iyul", "avgust", "sentabr", "oktabr", "noyabr", "dekabr",
)


@dataclass(frozen=True)
class _Labels:
    """Hisobotdagi barcha statik matnlar — krill yoki lotin alifbosida.

    Erkin matn (test/sessiya nomi, viloyat nomi) bu yerga kirmaydi; ular
    alohida ishlanadi (krillda translatsiya, lotinda DB qiymati o'zicha).
    """

    months: tuple
    malumot: str
    yil: str
    soat: str
    smena_suffix: str    # "-смена" / "-smena"
    smena_word: str      # nomsiz smena uchun fallback
    smenada_suffix: str  # "-сменада" / "-smenada"
    kuni: str
    kunlari: str
    tr: str
    hudud: str
    c_base: str
    participated: str
    not_participated: str
    shu_jumladan: str
    absent: str
    cheating_total: str
    soni: str
    pct: str
    at_entry: str
    during_test: str
    jami: str
    title_suffix: str    # sessiya/test nomidan keyin qo'shiladigan rasmiy matn


_CYRILLIC_LABELS = _Labels(
    months=_MONTHS_RU,
    malumot="МАЪЛУМОТ",
    yil="йил",
    soat="соат",
    smena_suffix="-смена",
    smena_word="смена",
    smenada_suffix="-сменада",
    kuni="куни",
    kunlari="кунлари",
    tr="Т/р",
    hudud="Ҳудуд номи",
    c_base="Иштирок этадиган абитуриентлар сони",
    participated="Тест синовларида иштирок этганлар",
    not_participated="Иштирок этмаганлар",
    shu_jumladan="шу жумладан:",
    absent="Тест синовларида қатнашмаганлар сони",
    cheating_total="Четлатилганлар сони",
    soni="сони",
    pct="%",
    at_entry="ҳудудга киришда",
    during_test="тест жараёнида",
    jami="Жами:",
    title_suffix="бўйича тест синовларида абитуриентлар иштироки тўғрисида",
)

_LATIN_LABELS = _Labels(
    months=_MONTHS_LATIN,
    malumot="MA'LUMOT",
    yil="yil",
    soat="soat",
    smena_suffix="-smena",
    smena_word="smena",
    smenada_suffix="-smenada",
    kuni="kuni",
    kunlari="kunlari",
    tr="T/r",
    hudud="Hudud nomi",
    c_base="Ishtirok etadigan abituriyentlar soni",
    participated="Test sinovlarida ishtirok etganlar",
    not_participated="Ishtirok etmaganlar",
    shu_jumladan="shu jumladan:",
    absent="Test sinovlarida qatnashmaganlar soni",
    cheating_total="Chetlatilganlar soni",
    soni="soni",
    pct="%",
    at_entry="hududga kirishda",
    during_test="test jarayonida",
    jami="Jami:",
    title_suffix="bo'yicha test sinovlarida abituriyentlar ishtiroki to'g'risida",
)

# === Ranglar (shablon theme ranglaridan hisoblangan) ===
_HEADER_FILL = "C5E0B4"   # accent6 (yashil) + tint 0.6 — sarlavha foni
_TOTAL_FONT = "1F4E79"    # accent1 (ko'k) + tint -0.5 — "Жами" matni
_RED = "C00000"           # "Иштирок этмаганлар" sonini ajratish uchun
_FONT_NAME = "Arial"

# Ustun kengliklari (shablondan)
_COL_WIDTHS = {
    "A": 13.0, "B": 74.0, "C": 46.0, "D": 25.0, "E": 15.0,
    "F": 28.0, "G": 18.0, "H": 36.0, "I": 25.0, "J": 25.0,
}

# Mingliklar ajratgichi — probel (masalan 1 292). Excelda oddiy sonning ajratgich
# belgisi (#,##0 dagi ","), lokal prefiksidan ("[$-...]") qat'i nazar, HAR DOIM
# Windows mintaqaviy sozlamasidan olinadi — shuning uchun uni majburlab bo'lmaydi.
# Yechim: magnitude shartlari bilan probelni LITERAL sifatida qo'yamiz — bu har
# qanday lokalda bir xil ishlaydi (999 999 999 gacha to'g'ri guruhlaydi):
#   >= 1 000 000 → "1 234 567"; >= 1 000 → "12 345"; qolgani → "292".
_NUM_FMT = '[>=1000000]#" "###" "##0;[>=1000]#" "##0;0'
_PCT_FMT = "0.0"

# Viloyat nomlarini krill alifbosiga o'tkazish — DB lotin yozuvida saqlaydi,
# hisobot esa (sarlavhalar kabi) krillda bo'lishi kerak. 14 ta viloyat barqaror
# to'plam; nom (apostrof variantlaridan qat'i nazar) bo'yicha moslaymiz.
_REGION_CYRILLIC = {
    "qoraqalpogiston respublikasi": "Қорақалпоғистон Республикаси",
    "andijon viloyati": "Андижон вилояти",
    "namangan viloyati": "Наманган вилояти",
    "fargona viloyati": "Фарғона вилояти",
    "buxoro viloyati": "Бухоро вилояти",
    "xorazm viloyati": "Хоразм вилояти",
    "surxondaryo viloyati": "Сурхондарё вилояти",
    "qashqadaryo viloyati": "Қашқадарё вилояти",
    "jizzax viloyati": "Жиззах вилояти",
    "navoiy viloyati": "Навоий вилояти",
    "samarqand viloyati": "Самарқанд вилояти",
    "sirdaryo viloyati": "Сирдарё вилояти",
    "toshkent viloyati": "Тошкент вилояти",
    "toshkent shahri": "Тошкент шаҳри",
}


def _region_cyrillic(name: str) -> str:
    """Lotin viloyat nomini krill ekvivalentiga o'tkazadi.

    Apostrof variantlari (ʻ ' ‘ ’ `) va katta-kichik harf farqi normallashtiriladi.
    Rasmiy 14 ta viloyat uchun aniq moslik ishlatiladi; topilmasa — umumiy
    transliterator orqali o'tkaziladi.
    """
    key = name.strip().lower()
    for ch in ("ʻ", "'", "‘", "’", "`", "ʼ"):
        key = key.replace(ch, "")
    mapped = _REGION_CYRILLIC.get(key)
    return mapped if mapped is not None else _latin_to_cyrillic(name)


# === Umumiy O'zbek lotin → krill transliteratori ===
# Erkin matn (test/sessiya nomlari) uchun. Apostrof variantlari avval bitta
# `'` ga normallashtiriladi; o'/g' modifikatorlari va tutuq belgisi (ъ) farqlanadi.
_APOSTROPHES = ("ʻ", "‘", "’", "`", "ʼ", "ʹ")

_DIGRAPHS = {
    "o'": "ў", "g'": "ғ",
    "sh": "ш", "ch": "ч",
    "yo": "ё", "yu": "ю", "ya": "я", "ye": "е", "ts": "ц",
}
_SINGLES = {
    "a": "а", "b": "б", "c": "с", "d": "д", "f": "ф", "g": "г",
    "h": "ҳ", "i": "и", "j": "ж", "k": "к", "l": "л", "m": "м",
    "n": "н", "o": "о", "p": "п", "q": "қ", "r": "р", "s": "с",
    "t": "т", "u": "у", "v": "в", "w": "в", "x": "х", "y": "й", "z": "з",
}


def _latin_to_cyrillic(text: str) -> str:
    """O'zbek lotin yozuvini krillga o'tkazadi (katta-kichik harf saqlanadi).

    Allaqachon krill bo'lgan belgilar va boshqa simvollar o'zgarmaydi.
    """
    s = text
    for ap in _APOSTROPHES:
        s = s.replace(ap, "'")

    out: list[str] = []
    i = 0
    n = len(s)
    while i < n:
        ch = s[i]
        two = s[i:i + 2].lower()
        if two in _DIGRAPHS:
            cyr = _DIGRAPHS[two]
            out.append(cyr.upper() if ch.isupper() else cyr)
            i += 2
            continue
        low = ch.lower()
        if low == "'":
            # o'/g' yuqorida ushlangan — qolgan apostrof = tutuq belgisi
            out.append("ъ")
            i += 1
            continue
        if low == "e":
            # so'z boshida → э, aks holda → е
            prev = s[i - 1] if i > 0 else ""
            initial = (i == 0) or (not prev.isalpha() and prev != "'")
            cyr = "э" if initial else "е"
            out.append(cyr.upper() if ch.isupper() else cyr)
            i += 1
            continue
        if low in _SINGLES:
            cyr = _SINGLES[low]
            out.append(cyr.upper() if ch.isupper() else cyr)
            i += 1
            continue
        out.append(ch)
        i += 1
    return "".join(out)

_thin = Side(style="thin", color="000000")
_medium = Side(style="medium", color="000000")


def _format_date(d: date, labels: _Labels) -> str:
    """date → "2026 йил 10 июль" / "2026 yil 10 iyul" ko'rinishida."""
    return f"{d.year} {labels.yil} {d.day} {labels.months[d.month]}"


def _format_date_range(d1: date, d2: date, labels: _Labels) -> str:
    """Sana oralig'i → "2026 йил 10-17 июль" (bir oy/yil bo'lsa) ko'rinishida."""
    if d1 == d2:
        return _format_date(d1, labels)
    if d1.year == d2.year and d1.month == d2.month:
        return f"{d1.year} {labels.yil} {d1.day}-{d2.day} {labels.months[d2.month]}"
    if d1.year == d2.year:
        return (
            f"{d1.year} {labels.yil} {d1.day} {labels.months[d1.month]} - "
            f"{d2.day} {labels.months[d2.month]}"
        )
    return f"{_format_date(d1, labels)} - {_format_date(d2, labels)}"


def _format_day_month(d: date, labels: _Labels) -> str:
    """date → "10 июль" / "10 iyul" (kun + oy, yilsiz)."""
    return f"{d.day} {labels.months[d.month]}"


def _format_day_month_range(d1: date, d2: date, labels: _Labels) -> str:
    """Sana oralig'i → "14-23 июль" (yilsiz) ko'rinishida."""
    if d1 == d2:
        return _format_day_month(d1, labels)
    if d1.month == d2.month and d1.year == d2.year:
        return f"{d1.day}-{d2.day} {labels.months[d2.month]}"
    return f"{_format_day_month(d1, labels)} - {_format_day_month(d2, labels)}"


def _scope_info_line(
    stats: DashboardStatsResponse,
    now: datetime,
    day_from: date | None,
    day_to: date | None,
    labels: _Labels,
) -> str:
    """3-qator uchun o'ng burchakdagi sana/vaqt/smena yozuvi."""
    if stats.scope == "smena" and stats.day is not None:
        smena = (
            f"{stats.smena_number}{labels.smena_suffix}"
            if stats.smena_number is not None
            else (stats.smena_name or labels.smena_word)
        )
        return f"{_format_date(stats.day, labels)}, {labels.soat} {now:%H:%M}, {smena}"
    if stats.scope == "day" and stats.day is not None:
        return _format_date(stats.day, labels)
    if stats.scope == "overall" and day_from is not None and day_to is not None:
        return _format_date_range(day_from, day_to, labels)
    return f"{labels.soat} {now:%H:%M}"


def _c_header_text(
    stats: DashboardStatsResponse,
    day_from: date | None,
    day_to: date | None,
    labels: _Labels,
) -> str:
    """C ustuni sarlavhasi — scope'ga qarab prefiks bilan.

    smena   → "1-сменада иштирок этадиган абитуриентлар сони"
    kunlik  → "10 июль куни иштирок этадиган абитуриентлар сони"
    umumiy  → "14-23 июль кунлари иштирок этадиган абитуриентлар сони"
    """
    base = labels.c_base
    # Prefiksdan keyin qator uziladi (Excel'da ALT+Enter = "\n", wrap_text yoqilgan)
    prefix = ""
    if stats.scope == "smena" and stats.smena_number is not None:
        prefix = f"{stats.smena_number}{labels.smenada_suffix}\n"
    elif stats.scope == "day" and stats.day is not None:
        prefix = f"{_format_day_month(stats.day, labels)} {labels.kuni}\n"
    elif stats.scope == "overall" and day_from is not None and day_to is not None:
        prefix = f"{_format_day_month_range(day_from, day_to, labels)} {labels.kunlari}\n"
    if prefix:
        return prefix + base[0].lower() + base[1:]
    return base


def _border(*, left: Side = _thin, right: Side = _thin,
            top: Side = _thin, bottom: Side = _thin) -> Border:
    return Border(left=left, right=right, top=top, bottom=bottom)


def build_session_stats_excel(
    stats: DashboardStatsResponse,
    *,
    title: str,
    day_from: date | None = None,
    day_to: date | None = None,
    latin: bool = False,
) -> bytes:
    """Dashboard statistikasidan rasmiy "МАЪЛУМОТ" Excel hisobotini yig'adi.

    `title` — hisobot tepasidagi sarlavha (test sessiya nomi + test nomi).
    `day_from`/`day_to` — sessiyaning kun oralig'i (umumiy/kunlik sarlavhalar uchun).
    `latin` — True bo'lsa hisobot o'zbek lotin alifbosida, aks holda krillda.
    """
    now = datetime.now(UZ_TZ)
    labels = _LATIN_LABELS if latin else _CYRILLIC_LABELS

    wb = Workbook()
    ws = wb.active
    ws.title = labels.malumot
    ws.sheet_view.showGridLines = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    last_col = 10  # A..J
    last_letter = get_column_letter(last_col)

    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # --- 1-qator: sessiya/test nomi + rasmiy izoh ("... бўйича тест синовларида
    # абитуриентлар иштироки тўғрисида"). Krillda nom translatsiya qilinadi,
    # lotinda DB qiymati o'zicha ishlatiladi. ---
    base_title = title if latin else _latin_to_cyrillic(title)
    ws.merge_cells(f"A1:{last_letter}1")
    c = ws["A1"]
    # Test nomi / rasmiy izoh / МАЪЛУМОТ — har biri ALT+ENTER (\n) bilan alohida
    # qatorda, bitta katakda
    c.value = f"{base_title}\n{labels.title_suffix}\n{labels.malumot}"
    c.font = Font(name=_FONT_NAME, size=26, bold=True)
    c.alignment = center
    ws.row_dimensions[1].height = 120

    # --- 2-qator: o'ng burchakda sana/vaqt/smena ---
    ws.merge_cells(f"H2:{last_letter}2")
    c = ws["H2"]
    c.value = _scope_info_line(stats, now, day_from, day_to, labels)
    c.font = Font(name=_FONT_NAME, size=16, bold=True)
    c.alignment = right
    ws.row_dimensions[2].height = 24

    # --- 3-5-qator: jadval sarlavhalari ---
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    header_font = Font(name=_FONT_NAME, size=20, bold=True)

    merges = {
        "A3:A5": labels.tr,
        "B3:B5": labels.hudud,
        "C3:C5": _c_header_text(stats, day_from, day_to, labels),
        "D3:E4": labels.participated,
        "F3:G4": labels.not_participated,
        "H3:J3": labels.shu_jumladan,
        "H4:H5": labels.absent,
        "I4:J4": labels.cheating_total,
    }
    singles = {
        "D5": labels.soni, "E5": labels.pct,
        "F5": labels.soni, "G5": labels.pct,
        "I5": labels.at_entry, "J5": labels.during_test,
    }
    for rng, text in merges.items():
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = text
    for coord, text in singles.items():
        ws[coord].value = text

    # Sarlavha kataklariga stil (merge ichidagi barcha kataklarga ham border)
    for row in range(3, 6):
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.alignment = center
            cell.fill = header_fill
            cell.border = _border()
    ws.row_dimensions[3].height = 40
    ws.row_dimensions[4].height = 52
    ws.row_dimensions[5].height = 66

    # --- 6+ : viloyatlar bo'yicha qatorlar ---
    data_font = Font(name=_FONT_NAME, size=24)
    red_font = Font(name=_FONT_NAME, size=24, color=_RED)
    bold_font = Font(name=_FONT_NAME, size=24, bold=True)

    first_data_row = 6
    row = first_data_row
    for idx, region in enumerate(stats.regions, start=1):
        st = region.stats
        total = st.total.total
        cheating_total = st.cheating.total
        at_entry = st.cheating.at_entry
        during_test = st.cheating.during_test
        # F = иштирок этмаганлар = umuman kelmaganlar + chetlatilganlar
        not_participated = st.not_attended.total + cheating_total

        # B ustuni: lotinda DB nomi o'zicha, krillda translatsiya qilingan nom
        region_name = (
            region.region_name if latin else _region_cyrillic(region.region_name)
        )
        ws.cell(row=row, column=1, value=idx)                 # A: T/r
        ws.cell(row=row, column=2, value=region_name)         # B: hudud
        ws.cell(row=row, column=3, value=total)               # C: total
        ws.cell(row=row, column=4, value=f"=C{row}-F{row}")   # D: ishtirok etgan soni
        ws.cell(row=row, column=5,
                value=f"=IFERROR(D{row}*100/C{row},0)")       # E: %
        ws.cell(row=row, column=6, value=not_participated)    # F: ishtirok etmagan soni
        ws.cell(row=row, column=7,
                value=f"=IFERROR(F{row}*100/C{row},0)")       # G: %
        ws.cell(row=row, column=8,
                value=f"=C{row}-D{row}-I{row}-J{row}")        # H: qatnashmaganlar
        ws.cell(row=row, column=9, value=at_entry)            # I: kirishda chetlatilgan
        ws.cell(row=row, column=10, value=during_test)        # J: testda chetlatilgan

        # Stil
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = _border()
            if col == 2:
                cell.alignment = left
            else:
                cell.alignment = center
            if col == 3:
                cell.font = bold_font
            elif col == 6:
                cell.font = red_font
            else:
                cell.font = data_font
        for col in (3, 4, 6, 8):
            ws.cell(row=row, column=col).number_format = _NUM_FMT
        for col in (5, 7):
            ws.cell(row=row, column=col).number_format = _PCT_FMT
        ws.row_dimensions[row].height = 44
        row += 1

    last_data_row = row - 1

    # --- Жами: yig'indi qatori ---
    total_row = row
    ws.merge_cells(f"A{total_row}:B{total_row}")
    ws.cell(row=total_row, column=1, value=labels.jami)

    if last_data_row >= first_data_row:

        def _sum(letter: str) -> str:
            return f"=SUM({letter}{first_data_row}:{letter}{last_data_row})"

        ws.cell(row=total_row, column=3, value=_sum("C"))
        ws.cell(row=total_row, column=4, value=_sum("D"))
        ws.cell(row=total_row, column=5,
                value=f"=IFERROR(D{total_row}*100/C{total_row},0)")
        ws.cell(row=total_row, column=6, value=_sum("F"))
        ws.cell(row=total_row, column=7,
                value=f"=IFERROR(F{total_row}*100/C{total_row},0)")
        ws.cell(row=total_row, column=8, value=_sum("H"))
        ws.cell(row=total_row, column=9, value=_sum("I"))
        ws.cell(row=total_row, column=10, value=_sum("J"))

    total_font = Font(name=_FONT_NAME, size=26, bold=True, color=_TOTAL_FONT)
    for col in range(1, last_col + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.font = total_font
        cell.alignment = center if col != 1 else Alignment(
            horizontal="center", vertical="center"
        )
        cell.border = _border(top=_medium, bottom=_medium)
    for col in (3, 4, 6, 8, 9, 10):
        ws.cell(row=total_row, column=col).number_format = _NUM_FMT
    for col in (5, 7):
        ws.cell(row=total_row, column=col).number_format = _PCT_FMT
    ws.row_dimensions[total_row].height = 46

    # Tashqi medium ramka (chap/o'ng) — jadval bo'ylab
    for r in range(3, total_row + 1):
        lc = ws.cell(row=r, column=1)
        rc = ws.cell(row=r, column=last_col)
        lb = lc.border
        rb = rc.border
        lc.border = Border(left=_medium, right=lb.right, top=lb.top, bottom=lb.bottom)
        rc.border = Border(left=rb.left, right=_medium, top=rb.top, bottom=rb.bottom)
    # Yuqori medium chiziq (sarlavha tepasi)
    for col in range(1, last_col + 1):
        cell = ws.cell(row=3, column=col)
        b = cell.border
        cell.border = Border(left=b.left, right=b.right, top=_medium, bottom=b.bottom)

    ws.print_title_rows = "4:6"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
