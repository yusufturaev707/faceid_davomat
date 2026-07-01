"""Admin statistika — talabgorlar holati (kelmagan + chetlatilgan) Excel.

Tanlangan kontekstda (smena / kun / umumiy) talabgorlarni status bo'yicha
guruhlab .xlsx hosil qiladi. Guruhlar tartibi (Status ustuni qiymatlari):

  1) "Kelmadi"                     — is_entered=False, is_cheating=False, is_applied=False
  2) "Kirishda chetlatildi"        — is_cheating=True, reason_type.key=1
  3) "Test jarayonida chetlatildi" — is_cheating=True, reason_type.key=2
  4) "Chetlatildi"                 — is_cheating=True, boshqa/aniqlanmagan reason_type

Har guruh ichida qatorlar: sana → region → zone → smena → guruh tartibida.
Chetlatilganlar uchun "Sabab" ustunida chetlatish sababi (Reason.name)
ko'rsatiladi.

`davomat_bot_absentees.build_absentees_excel` (Telegram bot) dan farqli — bu
hisobot admin statistika sahifasi uchun, chetlatilganlarni ham qamrab oladi.
Smena hal qilish mantig'i shu moduldan qayta ishlatiladi.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, or_, select
from sqlalchemy.orm import Session

from app.models.cheating_log import CheatingLog
from app.models.reason import Reason
from app.models.reason_type import ReasonType
from app.models.region import Region
from app.models.smena import Smena
from app.models.student import Student
from app.models.test_session_smena import TestSessionSmena
from app.models.zone import Zone
from app.services.davomat_bot_absentees import AbsenteesError, _resolve_smenas
from app.services.session_dashboard_stats import (
    ENTRY_REASON_TYPE_KEY,
    TEST_REASON_TYPE_KEY,
)

# Excel uslublari — davomat bot hisoboti bilan bir xil yashil palitra.
_TITLE_FILL = PatternFill("solid", fgColor="1B5E20")
_META_FILL = PatternFill("solid", fgColor="2E7D32")
_HEADER_FILL = PatternFill("solid", fgColor="2E7D32")
_ZEBRA_FILL = PatternFill("solid", fgColor="E8F5E9")

_THIN = Side(style="thin", color="A5D6A7")
_THICK = Side(style="medium", color="1B5E20")
_BORDER_CELL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_HEADER = Border(left=_THIN, right=_THIN, top=_THICK, bottom=_THICK)

# Status guruhlari — (tartib indeksi, ko'rsatiladigan matn, shrift rangi).
# Tartib indeksi qatorlarni saralashda birlamchi kalit: kelmadi → kirishda →
# test jarayonida → boshqa.
_STATUS_NOT_ENTERED = (0, "Kelmadi", "EF6C00")
_STATUS_ENTRY = (1, "Kirishda chetlatildi", "C62828")
_STATUS_TEST = (2, "Test jarayonida chetlatildi", "B71C1C")
_STATUS_OTHER = (3, "Chetlatildi", "C62828")


def _safe(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _status_for(is_cheating: bool, reason_type_key: int | None) -> tuple[int, str, str]:
    """(group, matn, rang) — is_cheating va reason_type.key bo'yicha."""
    if not is_cheating:
        return _STATUS_NOT_ENTERED
    if reason_type_key == ENTRY_REASON_TYPE_KEY:
        return _STATUS_ENTRY
    if reason_type_key == TEST_REASON_TYPE_KEY:
        return _STATUS_TEST
    return _STATUS_OTHER


def _fetch_rows(
    db: Session, *, session_smena_ids: list[int], allowed_region_ids: set[int]
) -> list[dict]:
    """Kelmagan VA chetlatilgan talabgorlarni (sabab bilan) qaytaradi.

    Qamrab olinadi:
      - chetlatilganlar (is_cheating=True), yoki
      - umuman kelmaganlar (is_entered=False AND is_applied=False).

    Saralash (Python tarafida): status guruhi → sana → region → zone → smena →
    guruh → joy.
    """
    rows = db.execute(
        select(
            Student.id,
            Student.last_name,
            Student.first_name,
            Student.middle_name,
            Student.imei,
            Student.gr_n,
            Student.sp_n,
            Student.is_cheating,
            TestSessionSmena.day.label("test_day"),
            TestSessionSmena.number.label("smena_number"),
            Smena.name.label("smena_name"),
            Zone.name.label("zone_name"),
            Zone.number.label("zone_number"),
            Region.name.label("region_name"),
            Region.number.label("region_number"),
            Reason.name.label("reason_name"),
            ReasonType.key.label("reason_type_key"),
        )
        .select_from(Student)
        .join(TestSessionSmena, TestSessionSmena.id == Student.session_smena_id)
        .join(Smena, Smena.id == TestSessionSmena.test_smena_id)
        .join(Zone, Zone.id == Student.zone_id)
        .join(Region, Region.id == Zone.region_id)
        .outerjoin(CheatingLog, CheatingLog.student_id == Student.id)
        .outerjoin(Reason, Reason.id == CheatingLog.reason_id)
        .outerjoin(ReasonType, ReasonType.id == Reason.reason_type_id)
        .where(
            Student.session_smena_id.in_(session_smena_ids),
            Region.id.in_(allowed_region_ids),
            or_(
                Student.is_cheating.is_(True),
                and_(
                    Student.is_entered.is_(False),
                    Student.is_applied.is_(False),
                ),
            ),
        )
    ).all()

    result: list[dict] = []
    for r in rows:
        is_cheating = bool(r.is_cheating)
        group, status_text, status_color = _status_for(
            is_cheating, r.reason_type_key
        )
        fio = " ".join(
            p for p in [r.last_name, r.first_name, r.middle_name] if p
        ).strip()
        result.append(
            {
                "group": group,
                "status_text": status_text,
                "status_color": status_color,
                "reason": (r.reason_name or "") if is_cheating else "",
                "fio": fio,
                "imei": r.imei or "",
                "gr_n": int(r.gr_n or 0),
                "sp_n": int(r.sp_n or 0),
                "region_name": r.region_name or "",
                "region_number": int(r.region_number or 0),
                "zone_name": r.zone_name or "",
                "zone_number": int(r.zone_number or 0),
                "test_day": r.test_day,
                "smena_number": int(r.smena_number or 0),
                "smena_name": r.smena_name or "",
            }
        )

    # Birlamchi: status guruhi; so'ng sana → region → zone → smena → guruh → joy.
    result.sort(
        key=lambda d: (
            d["group"],
            d["test_day"] or date.min,
            d["region_number"],
            d["region_name"],
            d["zone_number"],
            d["zone_name"],
            d["smena_number"],
            d["gr_n"],
            d["sp_n"],
        )
    )
    return result


def build_session_status_excel(
    db: Session,
    *,
    session_id: int,
    session_smena_id: int | None = None,
    test_day: date | None = None,
    allowed_region_ids: set[int],
) -> tuple[bytes, str, int]:
    """Talabgorlar holati ro'yxatini .xlsx bytes shaklida qaytaradi.

    Kontekst `build_absentees_excel` bilan bir xil aniqlanadi:
      - `session_smena_id` berilsa: bitta smena.
      - `test_day` berilsa: shu kunning barcha aktiv smenalari.
      - Ikkalasi `None` bo'lsa: sessiyaning barcha aktiv smenalari.

    Returns:
        (bytes, filename, row_count)
    """
    if not allowed_region_ids:
        raise AbsenteesError("Viloyatlar topilmadi")

    (
        test_session,
        smena_ids,
        resolved_day,
        smena_name,
        smena_number,
        scope,
    ) = _resolve_smenas(
        db,
        session_id=session_id,
        session_smena_id=session_smena_id,
        test_day=test_day,
    )
    test_name = (
        test_session.test.name if test_session.test else None
    ) or test_session.name

    rows = _fetch_rows(
        db,
        session_smena_ids=smena_ids,
        allowed_region_ids=allowed_region_ids,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Talabgorlar holati"

    columns = [
        ("№", 6),
        ("Sana", 13),
        ("Smena", 18),
        ("Viloyat", 24),
        ("Bino", 24),
        ("FIO", 52),
        ("IMEI (JShShIR)", 18),
        ("Gr", 8),
        ("Status", 26),
        ("Sabab", 42),
    ]
    n_cols = len(columns)
    last_col = get_column_letter(n_cols)
    col_status = 9
    col_reason = 10
    col_imei = 7
    col_fio = 6
    col_gr = 8

    # ── Sarlavha bloki ──────────────────────────────────────────────────
    if scope == "day":
        title_text = (
            f"TALABGORLAR HOLATI — {test_name} • "
            f"{resolved_day.isoformat() if resolved_day else '—'}"
        )
    elif scope == "total":
        title_text = f"TALABGORLAR HOLATI (UMUMIY) — {test_name}"
    else:
        title_text = f"TALABGORLAR HOLATI — {test_name}"

    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[1].height = 30

    # 2-qator: sana
    ws.merge_cells(f"A2:{last_col}2")
    date_cell = ws["A2"]
    if scope == "total":
        date_cell.value = "Sana: Barcha kunlar"
    else:
        date_cell.value = (
            f"Test sanasi: {resolved_day.isoformat() if resolved_day else '—'}"
        )
    date_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    date_cell.fill = _META_FILL
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # 3-qator: smena
    ws.merge_cells(f"A3:{last_col}3")
    smena_cell = ws["A3"]
    if scope == "smena":
        smena_label = smena_name or "—"
        if smena_number:
            smena_label = f"#{smena_number} • {smena_label}"
        smena_cell.value = f"Smena: {smena_label}"
    elif scope == "day":
        smena_cell.value = "Smena: Shu sanadagi barcha smenalar"
    else:
        smena_cell.value = "Smena: Barcha smenalar"
    smena_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    smena_cell.fill = _META_FILL
    smena_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    # 4-qator: status bo'yicha umumiy hisob
    not_entered = sum(1 for r in rows if r["group"] == _STATUS_NOT_ENTERED[0])
    entry = sum(1 for r in rows if r["group"] == _STATUS_ENTRY[0])
    test = sum(1 for r in rows if r["group"] == _STATUS_TEST[0])
    other = sum(1 for r in rows if r["group"] == _STATUS_OTHER[0])
    summary_parts = [f"Kelmadi: {not_entered}"]
    summary_parts.append(f"Kirishda chetlatildi: {entry}")
    summary_parts.append(f"Test jarayonida chetlatildi: {test}")
    if other:
        summary_parts.append(f"Boshqa chetlatish: {other}")
    summary_parts.append(f"Jami: {len(rows)}")

    ws.merge_cells(f"A4:{last_col}4")
    total_cell = ws["A4"]
    total_cell.value = " • ".join(summary_parts)
    total_cell.font = Font(
        name="Calibri", size=11, bold=True, italic=True, color="1B5E20"
    )
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 20

    ws.row_dimensions[5].height = 6

    # ── Jadval sarlavhalari (6-qator) ───────────────────────────────────
    header_row = 6
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER_HEADER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 26

    # ── Body ────────────────────────────────────────────────────────────
    data_start = header_row + 1
    for i, r in enumerate(rows, start=1):
        row_idx = data_start + i - 1
        day_str = (
            r["test_day"].isoformat()
            if isinstance(r.get("test_day"), date)
            else _safe(r.get("test_day"))
        )
        sm_n = r.get("smena_number") or 0
        sm_label = r.get("smena_name") or ""
        smena_disp = f"#{sm_n} • {sm_label}" if sm_n else sm_label
        values = [
            i,
            day_str,
            smena_disp,
            _safe(r["region_name"]),
            _safe(r["zone_name"]),
            _safe(r["fio"]),
            _safe(r["imei"]),
            r["gr_n"] or "",
            r["status_text"],
            _safe(r["reason"]),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=11)
            cell.border = _BORDER_CELL
            if col_idx == col_imei:
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == col_fio:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=False
                )
            elif col_idx == col_status:
                cell.font = Font(
                    name="Calibri", size=11, bold=True, color=r["status_color"]
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == col_reason:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            elif col_idx in (1, col_gr):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            if i % 2 == 0:
                cell.fill = _ZEBRA_FILL
        ws.row_dimensions[row_idx].height = 20

    if not rows:
        empty_row = data_start
        ws.merge_cells(
            start_row=empty_row, start_column=1, end_row=empty_row, end_column=n_cols
        )
        cell = ws.cell(
            row=empty_row,
            column=1,
            value="✓ Bu kontekstda kelmagan yoki chetlatilgan talabgor yo'q",
        )
        cell.font = Font(name="Calibri", size=12, italic=True, color="2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER_CELL
        ws.row_dimensions[empty_row].height = 28

    ws.freeze_panes = ws.cell(row=data_start, column=1)
    ws.sheet_view.showGridLines = False
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # ── Fayl baytlari ───────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    if scope == "day":
        safe_day = resolved_day.isoformat() if resolved_day else "no-date"
        filename = f"talabgorlar_holati_{safe_day}_umumiy.xlsx"
    elif scope == "total":
        filename = "talabgorlar_holati_umumiy.xlsx"
    else:
        safe_day = resolved_day.isoformat() if resolved_day else "no-date"
        filename = f"talabgorlar_holati_{safe_day}_smena_{smena_number or 'x'}.xlsx"
    return buffer.getvalue(), filename, len(rows)
