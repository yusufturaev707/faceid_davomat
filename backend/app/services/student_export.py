"""Talabalar ro'yxatini Excel (.xlsx) va PDF formatida eksport qilish.

Ustunlar (so'ralgan tartibda):
    Test, Hudud, Familiya, Ism, Otasining ismi, JSHSHIR,
    Seriya, Raqam, Sana, Smena, Guruh

Ikkala format ham chiroyli, bosib chiqarishga tayyor hisobot ko'rinishida:
sarlavha, meta qator (jami soni + yaratilgan vaqt), rangli sarlavhalar va
o'qishni osonlashtirish uchun navbatma-navbat (zebra) qatorlar.
"""

from __future__ import annotations

import os
from datetime import date, datetime
from io import BytesIO

# Hisobot ustunlari: (sarlavha, dict'dan qiymat oluvchi kalit)
# Qiymatlar `app.crud.student._student_to_dict` chiqaradigan dict'dan olinadi.
_COLUMNS: tuple[tuple[str, str], ...] = (
    ("Test", "test_name"),
    ("Hudud", "region_name"),
    ("Familiya", "last_name"),
    ("Ism", "first_name"),
    ("Otasining ismi", "middle_name"),
    ("JSHSHIR", "jshshir"),
    ("Seriya", "ps_ser"),
    ("Raqam", "ps_num"),
    ("Sana", "test_date"),
    ("Smena", "smena"),
    ("Guruh", "group"),
)

# Brend ranglari (frontend primary bilan uyg'un — ko'k)
_BRAND = "1E88E5"
_BRAND_DARK = "1565C0"
_HEADER_TEXT = "FFFFFF"
_ZEBRA = "EEF4FB"
_BORDER = "D6E2F0"


def _fmt_date(value: object) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y")
    return str(value or "")


def _row_values(r: dict) -> list[str]:
    """Bitta talaba dict'idan ustunlar tartibida string qiymatlar."""
    ps = r.get("ps_data") or {}
    src = {
        "test_name": r.get("test_name"),
        "region_name": r.get("region_name"),
        "last_name": r.get("last_name"),
        "first_name": r.get("first_name"),
        "middle_name": r.get("middle_name"),
        "jshshir": r.get("imei"),
        "ps_ser": ps.get("ps_ser"),
        "ps_num": ps.get("ps_num"),
        "test_date": _fmt_date(r.get("e_date")),
        "smena": r.get("smena_name"),
        "group": r.get("gr_n"),
    }
    out: list[str] = []
    for _, key in _COLUMNS:
        val = src.get(key)
        if key == "group" and (val in (None, 0)):
            val = ""  # 0 guruhni bo'sh ko'rsatamiz
        out.append("" if val is None else str(val))
    return out


# ============================ Excel (.xlsx) ============================


def build_students_xlsx(rows: list[dict], generated_at: datetime) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Talabalar"

    ncols = len(_COLUMNS) + 1  # + tartib raqami (№)
    thin = Side(style="thin", color=_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 1-qator: sarlavha (merge)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title = ws.cell(row=1, column=1, value="TALABALAR RO'YXATI")
    title.font = Font(bold=True, size=15, color=_HEADER_TEXT)
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.fill = PatternFill("solid", fgColor=_BRAND_DARK)
    ws.row_dimensions[1].height = 26

    # 2-qator: meta (jami + yaratilgan vaqt)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    meta = ws.cell(
        row=2,
        column=1,
        value=(
            f"Jami: {len(rows)} ta    |    "
            f"Yuklab olindi: {generated_at.strftime('%d.%m.%Y %H:%M')}"
        ),
    )
    meta.font = Font(italic=True, size=10, color="5B6B7B")
    meta.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 18

    # 3-qator: ustun sarlavhalari
    header_row = 3
    header_fill = PatternFill("solid", fgColor=_BRAND)
    header_font = Font(bold=True, size=11, color=_HEADER_TEXT)
    headers = ["№", *[h for h, _ in _COLUMNS]]
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[header_row].height = 22

    # Ma'lumot qatorlari
    for i, r in enumerate(rows):
        excel_row = header_row + 1 + i
        values = [i + 1, *_row_values(r)]
        zebra = i % 2 == 1
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=excel_row, column=col_idx, value=val)
            c.border = border
            c.alignment = center if col_idx == 1 else left
            if zebra:
                c.fill = PatternFill("solid", fgColor=_ZEBRA)

    # Ustun kengliklari (№ + 11 ustun)
    widths = (5, 22, 18, 16, 14, 18, 18, 9, 12, 12, 10, 9)
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================ PDF ============================


def _register_pdf_font() -> tuple[str, str]:
    """Unicode (Lotin/Uzbek) belgilar uchun TTF shrift ro'yxatdan o'tkazadi.

    Tizimda mos shrift topilmasa — Helvetica (Latin-1) ga qaytadi.
    Qaytaradi: (regular_font_name, bold_font_name).
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = (
        (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        ),
        (
            "/usr/share/fonts/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
        ),
        (r"C:\Windows\Fonts\arial.ttf", r"C:\Windows\Fonts\arialbd.ttf"),
    )
    for regular_path, bold_path in candidates:
        if not os.path.exists(regular_path):
            continue
        try:
            pdfmetrics.registerFont(TTFont("ExportFont", regular_path))
            if os.path.exists(bold_path):
                pdfmetrics.registerFont(TTFont("ExportFont-Bold", bold_path))
            else:
                pdfmetrics.registerFont(TTFont("ExportFont-Bold", regular_path))
            return "ExportFont", "ExportFont-Bold"
        except Exception:
            continue
    return "Helvetica", "Helvetica-Bold"


def build_students_pdf(rows: list[dict], generated_at: datetime) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    font, font_bold = _register_pdf_font()
    brand = colors.HexColor("#" + _BRAND)
    brand_dark = colors.HexColor("#" + _BRAND_DARK)
    zebra = colors.HexColor("#" + _ZEBRA)
    border = colors.HexColor("#" + _BORDER)

    cell_style = ParagraphStyle(
        "cell", fontName=font, fontSize=7.5, leading=9, alignment=TA_LEFT
    )
    header_style = ParagraphStyle(
        "hcell", fontName=font_bold, fontSize=8, leading=10,
        alignment=TA_CENTER, textColor=colors.white,
    )
    title_style = ParagraphStyle(
        "title", fontName=font_bold, fontSize=15, leading=18,
        alignment=TA_CENTER, textColor=brand_dark,
    )
    meta_style = ParagraphStyle(
        "meta", fontName=font, fontSize=9, leading=11,
        alignment=TA_CENTER, textColor=colors.HexColor("#5B6B7B"),
    )

    headers = ["№", *[h for h, _ in _COLUMNS]]
    table_data: list[list] = [[Paragraph(h, header_style) for h in headers]]
    for i, r in enumerate(rows):
        cells = [Paragraph(str(i + 1), cell_style)]
        cells += [Paragraph(_html_escape(v), cell_style) for v in _row_values(r)]
        table_data.append(cells)

    # Ustun kengliklari (landscape A4 da, № + 11 ustun) — nisbiy
    page = landscape(A4)
    usable = page[0] - 16 * mm
    weights = (3, 11, 9, 9, 8, 7, 10, 5, 6, 7, 6, 5)
    total_w = sum(weights)
    col_widths = [usable * w / total_w for w in weights]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), brand),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),  # № ustuni
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.4, border),
        ("LINEBELOW", (0, 0), (-1, 0), 0.8, brand_dark),
    ]
    for i in range(len(rows)):
        if i % 2 == 1:
            style.append(("BACKGROUND", (0, i + 1), (-1, i + 1), zebra))
    table.setStyle(TableStyle(style))

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=10 * mm,
        title="Talabalar ro'yxati",
    )
    meta_text = (
        f"Jami: {len(rows)} ta&nbsp;&nbsp;|&nbsp;&nbsp;"
        f"Yuklab olindi: {generated_at.strftime('%d.%m.%Y %H:%M')}"
    )
    elements = [
        Paragraph("TALABALAR RO'YXATI", title_style),
        Spacer(1, 2 * mm),
        Paragraph(meta_text, meta_style),
        Spacer(1, 4 * mm),
        table,
    ]
    doc.build(elements)
    return buf.getvalue()


def _html_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    )
